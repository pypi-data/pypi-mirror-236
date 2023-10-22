from typing import List, Tuple, Dict, Optional, Union
from openpyxl.styles import Font, Alignment
from itertools import combinations
from itertools import permutations
import matplotlib.pyplot as plt
from gurobipy import *
import networkx as nx
import pandas as pd
import numpy as np
from pulp import *
import openpyxl
import os

# Deshabilitar todas las advertencias de tipo UserWarning a nivel global
warnings.filterwarnings("ignore", category=UserWarning)

class SecuenciacionProgramacionLineal():
    def __init__(self, tareas : dict, funcion_objetivo : str = 'Tiempo Total Procesamiento', tiempos_entrega : dict = None, importancia : dict = None, verbose = False):
        '''
        SecuenciacionProgramacionLineal(tareas: dict[str ,dict[float, ..., float]])\n
        Secuenciación de tareas por estación para optimizacion de demora maxima por medio de progrmación lineal\n
        Argumentos:\n
        tareas : dict [dict]
            diccionario de key:tareas value:diccionarios de estaciones con duraciones de cada tarea respectiva en la estacion, en caso de estar repetida la estacion en
            el proceso se debe agregar el nombre de la estacioneseguido de raya al piso y el numero de la priorida, se debe ingresar en el diccionario de cada tarea las 
            estaciones en el orden que van a ser procesados\n
        funcion_objetivo:\n
            funciones objetivo conla cual se buscara el optimo para dicho valor, puede escoger entre las siguientes.
            \n* tiempo total procesamiento
            \n* importancia
            \n* tiempo total bloqueo
            \n* tiempo promedio flujo
            \n* tiempos retraso
        tiempos_entrega:\n
            diciconario con tiempos de entrega maximos para ada tarea, en caso de solo especificar uno fijar un número muy grande para las otras tareas.
        importnacia:\n
            diciconario con tiempos de entrega maximos para ada tarea, en caso de solo especificar uno fijar un número muy grande para las otras tareas.
        verbose:\n
            Controla la vista de la solución del modelo con tiempos y tamaño de salto del gap en busca de la solución
        Ejemplo:\n
            tareas = {'T1' : { 'M1':5  , 'M2':7   , 'M3':7   },\n
                    'T2' : { 'M1_1':3 , 'M2_1':7 , 'M1_2':2  , 'M2_2':7 },\n
                    'T3' : { 'M3':5 , 'M4':8   , 'M2':9 },\n
                    'T4' : { 'M1':4   , 'M4':7   , 'M3':6   },\n
                    'T5' : { 'M2_1':5 , 'M1':6   , 'M2_2':7  , 'M3':2 },\n
                    'T6' : { 'M1':8   , 'M3':5   , 'M2':4   },\n
                    'T7' : { 'M3_1':9 , 'M4':2   , 'M1':5    , 'M3_2':5 }}\n
            importancia = {'T1':2, 'T2':3, 'T3':4, 'T4':1, 'T5':1, 'T6':50, 'T7':300}\n
            tiempos_entrega = {'T1':50, 'T2':50, 'T3':50, 'T4':50, 'T5':50, 'T6':50, 'T7':50} \n
            SecuanciacionPL = SecuenciacionProgramacionLineal(tareas = tareas, tiempos_entrega = tiempos_entrega, importancia = importancia, funcion_objetivo = 'tiempo total bloqueo', verbose = True)\n
            tiempos_resultado, resumen_tareas, resumen_estaciones, tiempo_procesamiento_minimo = SecuanciacionPL.ResumenResultados()\n
            print(tiempos_resultado)\n
            print(resumen_tareas)\n
            print(resumen_estaciones)\n
            print(tiempo_procesamiento_minimo)\n
            SecuanciacionPL.Formulacion()\n
            SecuanciacionPL.DiagramaGantt()
        '''
        self.modelo = LpProblem("modelo_secuanciacion_estaciones", sense=LpMinimize)
        self.M = 0
        self.tareas = tareas
        for tarea in tareas.keys():
            for estacion in self.tareas[tarea]:
                if not isinstance(self.tareas[tarea][estacion], (int, float)):
                    raise ValueError(f'Valor de diccionario no entero o decimal en la tarea {tarea} y estación {estacion}.')
                self.M += self.tareas[tarea][estacion] * 2
        self.M = int(self.M)
        self.CrearVariables()
        self.RestriccionesSolapeTiempos()
        self.RestriccionSecuencia()
        self.RestriccionesTiempoMinimo()
        self.RestriccionesTiempoMuerto()
        self.RestriccionesTiempoFlujoTarea()
        # Manejo de errores tiempso de entrega
        if isinstance(tiempos_entrega, dict):
            self.tiempos_entrega = tiempos_entrega
            if set(map(id, self.tiempos_entrega)) == set(map(id, self.tareas)):
                for tarea in self.tiempos_entrega.keys():
                    if not isinstance(self.tiempos_entrega[tarea], (int, float)):
                        raise ValueError(f'Valor de diccionario no entero o decimal en la tarea {tarea}.')
                self.RestriccionTiemposEntrega()
            else:
                raise ValueError(f'Llaves de diccionario no correspondientes a tamaño.')
        elif tiempos_entrega != None:
            raise ValueError(f'Variable tiempos de entrega no es un diccionario.')
        # Manejo de errores improtnacias
        if isinstance(importancia, dict):
            self.importancia = importancia
            if set(map(id, self.importancia)) == set(map(id, self.tareas)):
                for tarea in self.importancia.keys():
                    if not isinstance(self.importancia[tarea], (int, float)):
                        raise ValueError(f'Valor de diccionario no entero o decimal en la tarea {tarea}.')
            else:
                raise ValueError(f'Llaves de diccionario no correspondientes a tamaño.')
        elif importancia != None:
            raise ValueError(f'Variable tiempos de entrega no es un diccionario.')
        self.FuncionObjetivo(funcion_objetivo)
        self.Solucionar(verbose)
        self.ResumenResultados()

    # Crear varibales modelo
    def CrearVariables(self):
        self.nombres_tareas = list(self.tareas.keys())
        self.nombres_estaciones = list(set([estacion.split('_')[0] for tarea in self.tareas.keys() for estacion in self.tareas[tarea]]))
        #self.nombres_estaciones_tareas = dict([(tarea,list(self.tareas[tarea].keys())) for tarea in self.nombres_tareas]) CAMBIO POSIBLE
        self.nombres_estaciones_tareas = {tarea:list(self.tareas[tarea].keys()) for tarea in self.nombres_tareas}
        # Crear variables Tiempos de inicio
        self.nombres_tiempos_inicio = [tarea+'_'+estacion for tarea in self.nombres_tareas for estacion in self.nombres_estaciones_tareas[tarea]]
        self.tiempos_inicio = LpVariable.dicts("TiemposInicio", self.nombres_tiempos_inicio , lowBound=0, cat='Continuous')
        # Crear agrupacion por Tareas
        self.diccionario_nombres_secuencia = dict((tarea,[]) for tarea in self.nombres_tareas)
        for nombres_tiempos_inicio in self.nombres_tiempos_inicio:
            nombres_tiempos_inicio_split = nombres_tiempos_inicio.split('_')
            self.diccionario_nombres_secuencia[nombres_tiempos_inicio_split[0]].append(nombres_tiempos_inicio)
        # Crear agrupacion por Estaciones
        self.diccionario_nombres_estaciones = dict((estacion,[]) for estacion in self.nombres_estaciones)
        for estacion in self.nombres_estaciones:
            for nombres_tiempos_inicio in self.nombres_tiempos_inicio:
                if estacion in nombres_tiempos_inicio:
                    self.diccionario_nombres_estaciones[estacion].append(nombres_tiempos_inicio)
        # Crear variables Binarias de activacion
        self.nombres_binarias = []
        for estacion in self.nombres_estaciones:
            self.nombres_binarias += list(combinations(self.diccionario_nombres_estaciones[estacion],2))
        self.binaria_activacion = LpVariable.dicts("BinariaActivacion", self.nombres_binarias , cat='Binary')
    
    # Restriccion de Solape de tiempos en misma estacion
    def RestriccionesSolapeTiempos(self):
        for nombres_tiempos_inicio in self.nombres_binarias:
            nti1, nti2 = nombres_tiempos_inicio[0], nombres_tiempos_inicio[1] 
            tnti1, tnti2 = nti1.split('_',1)[0], nti2.split('_',1)[0]
            mnti1, mnti2 = nti1.split('_',1)[1], nti2.split('_',1)[1]
            dur1, dur2 = self.tareas[tnti1][mnti1], self.tareas[tnti2][mnti2]
            self.modelo += self.tiempos_inicio[nti1] + dur1 <= self.tiempos_inicio[nti2] + self.binaria_activacion[nti1, nti2]*self.M
            self.modelo += self.tiempos_inicio[nti2] + dur2 <= self.tiempos_inicio[nti1] + (1-self.binaria_activacion[nti1, nti2])*self.M

    # Restriccion de Secuencia en estacion de cada tarea
    def RestriccionSecuencia(self):
        for tarea in self.diccionario_nombres_secuencia.keys():
            for n in range(len(self.diccionario_nombres_secuencia[tarea])-1):
                nti1, nti2 = self.diccionario_nombres_secuencia[tarea][n], self.diccionario_nombres_secuencia[tarea][n+1]
                tnti1 = nti1.split('_',1)[0]
                mnti1 = nti1.split('_',1)[1]
                self.modelo += self.tiempos_inicio[nti1] + self.tareas[tnti1][mnti1] <= self.tiempos_inicio[nti2]

    # Restriccion de Tiempos entrega
    def RestriccionTiemposEntrega(self):
        self.tiempo_retraso = LpVariable.dicts("TiempoRetraso", self.nombres_tareas , lowBound=0, cat='Continuous')
        for tarea, nombres_estaciones in self.nombres_estaciones_tareas.items():
            ultima_estacion = nombres_estaciones[-1]
            self.modelo += self.tiempos_inicio[tarea+'_'+ultima_estacion] + self.tareas[tarea][ultima_estacion] - self.tiempo_retraso[tarea] <= self.tiempos_entrega[tarea]

    # Restriccion de Tiempo Minimo (MAKESPAN)
    def RestriccionesTiempoMinimo(self):
        self.tiempo_minimo = LpVariable("TiempoTotalProcesamiento", lowBound=0, cat='Continuous')
        for tarea in self.diccionario_nombres_secuencia.keys():
            ultima_estacion = self.diccionario_nombres_secuencia[tarea][-1]
            tnti, mnti = ultima_estacion.split('_',1)[0], ultima_estacion.split('_',1)[1] 
            self.modelo += self.tiempos_inicio[ultima_estacion] + self.tareas[tnti][mnti] <= self.tiempo_minimo

    # Restriccion de Tiempo Muerto
    def RestriccionesTiempoMuerto(self):
        self.tiempos_inicio_estacion = LpVariable.dicts("TiempoInicioEstacion", self.nombres_estaciones , lowBound=0, cat='Continuous')
        self.tiempos_fin_estacion = LpVariable.dicts("TiempoFinEstacion", self.nombres_estaciones , lowBound=0, cat='Continuous')
        for estacion, nombres_tareas in self.diccionario_nombres_estaciones.items():
            for tarea_estacion in nombres_tareas:
                self.modelo +=  self.tiempos_inicio_estacion[estacion] <= self.tiempos_inicio[tarea_estacion]
                self.modelo +=  self.tiempos_inicio[tarea_estacion] + self.tareas[tarea_estacion.split('_', 1)[0]][tarea_estacion.split('_', 1)[1]] <= self.tiempos_fin_estacion[estacion]

    # Restriccion de Tiempo promedio flujo
    def RestriccionesTiempoFlujoTarea(self):
        self.tiempos_inicio_tarea = LpVariable.dicts("TiempoInicioTarea", self.nombres_tareas , lowBound=0, cat='Continuous')
        self.tiempos_fin_tarea = LpVariable.dicts("TiempoFinTarea", self.nombres_tareas , lowBound=0, cat='Continuous')
        for tarea, nombres_estaciones in self.diccionario_nombres_secuencia.items():
            for tarea_estacion in nombres_estaciones:
                self.modelo +=  self.tiempos_inicio_tarea[tarea] <= self.tiempos_inicio[tarea_estacion]
                self.modelo +=  self.tiempos_inicio[tarea_estacion] + self.tareas[tarea_estacion.split('_', 1)[0]][tarea_estacion.split('_', 1)[1]] <= self.tiempos_fin_tarea[tarea]

    # Declaracion de Funcion objetivo
    def FuncionObjetivo(self, funcion):
        if funcion.lower() == 'tiempo total procesamiento':
            # Declaracion de Funcion Tiempo total procesamiento     
            self.modelo += self.tiempo_minimo
        elif funcion.lower() == 'importancia':
            # Declaracion de Funcion importancia
            if not hasattr(self, 'importacia'):
                raise Exception('No ingreso las importacias para calcular la función objetivo')
            objetivo = 0
            for tarea, nombres_estaciones in self.nombres_estaciones_tareas.items():
                ultima_estacion = nombres_estaciones[-1]
                objetivo += (self.tiempos_inicio[tarea+'_'+ultima_estacion] + self.tareas[tarea][ultima_estacion]) * self.importancia[tarea] 
            self.modelo += objetivo
        elif funcion.lower() == 'tiempo total bloqueo':
            # Declaracion de Funcion tiempo total bloqueo     
            objetivo = 0
            for estacion in self.nombres_estaciones:
                objetivo += self.tiempos_fin_estacion[estacion] - self.tiempos_inicio_estacion[estacion]
            for tarea in self.nombres_tareas:
                for estacion in self.tareas[tarea].keys():
                    objetivo -= self.tareas[tarea][estacion] 
            self.modelo += objetivo
        elif funcion.lower() == 'tiempo promedio flujo':
            # Declaracion de Funcion tiempo promedio flujo
            objetivo = 0
            for tarea in self.nombres_tareas:
                objetivo += self.tiempos_fin_tarea[tarea] - self.tiempos_inicio_tarea[tarea]
            self.modelo += (objetivo/len(self.nombres_tareas)) + (self.tiempo_minimo*(1/self.M))
        elif funcion.lower() == 'tiempos retraso':
            self.modelo += lpSum(self.tiempo_retraso[tarea] for tarea in self.nombres_tareas) 
        else:
            raise ValueError('Funcion objetivo no parametrizada')

    # Solucionar modelo
    def Solucionar(self, verbose):
        self.status = self.modelo.solve( solver = GUROBI(msg = verbose) )
        if LpStatus[self.modelo.status] == 'Optimal':
            self.horizonteTemporal = round(value(self.modelo.objective),0)
        else:
            raise ValueError('Modelo en estado: '+LpStatus[self.modelo.status])

    # Diccionario de tiempos de inicio
    def ResumenResultados(self):
        self.tiempos_resultado = {}
        self.resumen_tareas = pd.DataFrame(columns=['Tiempo inicio tarea','Tiempo fin tarea','Tiempo retraso'])
        self.resumen_estaciones = pd.DataFrame(columns=['Tiempo inicio estacion','Tiempo fin estacion'])
        for v in self.modelo.variables():
            if 'TiemposInicio' in str(v):
                nombre = str(v)
                nombre = nombre.replace('TiemposInicio_','')
                self.tiempos_resultado[nombre] = round(v.varValue,0)
            if 'TiempoInicioEstacion' in str(v):
                nombre = str(v).split('_')
                self.resumen_estaciones.loc[nombre[1], 'Tiempo inicio estacion'] = round(v.varValue,0)
            if 'TiempoFinEstacion' in str(v):
                nombre = str(v).split('_')
                self.resumen_estaciones.loc[nombre[1], 'Tiempo fin estacion'] = round(v.varValue,0)
            if 'TiempoInicioTarea' in str(v):
                nombre = str(v).split('_')
                self.resumen_tareas.loc[nombre[1], 'Tiempo inicio tarea'] = round(v.varValue,0)
            if 'TiempoFinTarea' in str(v):
                nombre = str(v).split('_')
                self.resumen_tareas.loc[nombre[1], 'Tiempo fin tarea'] = round(v.varValue,0)
            if 'TiempoRetraso' in str(v):
                nombre = str(v).split('_')
                self.resumen_tareas.loc[nombre[1], 'Tiempo retraso'] = round(v.varValue,0)
            if 'TiempoTotalProcesamiento' in str(v):
                self.tiempo_procesamiento_minimo = round(v.varValue,0)
        # Correccion de tiempos de incio tarea 
        for tarea in self.nombres_tareas:
            tareas_inicio = []
            for tarea_estacion in self.tiempos_resultado.keys():
                if tarea in tarea_estacion:
                    tareas_inicio.append(self.tiempos_resultado[tarea_estacion])
            self.resumen_tareas.loc[tarea, 'Tiempo inicio tarea'] = np.array(tareas_inicio).min()  
        # Correccion de tiempos de incio estación 
        for estacion in self.nombres_estaciones:
            estaciones_inicio = []
            for tarea_estacion in self.tiempos_resultado.keys():
                if estacion in tarea_estacion:
                    estaciones_inicio.append(self.tiempos_resultado[tarea_estacion])
            self.resumen_estaciones.loc[estacion, 'Tiempo inicio estacion'] = np.array(estaciones_inicio).min()  
        # Cálculo de métricas
        self.resumen_tareas['Tiempo flujo'] = self.resumen_tareas['Tiempo fin tarea'] - self.resumen_tareas['Tiempo inicio tarea']
        self.resumen_estaciones['Tiempo bloqueo'] = self.resumen_estaciones['Tiempo fin estacion'] - self.resumen_estaciones['Tiempo inicio estacion']
        for est in self.resumen_estaciones.index:
            for tarea in self.nombres_tareas:
                for estacion in self.tareas[tarea].keys():
                    if est in estacion:
                        self.resumen_estaciones.loc[est,'Tiempo bloqueo'] -= self.tareas[tarea][estacion]
        return self.tiempos_resultado, self.resumen_tareas, self.resumen_estaciones, self.tiempo_procesamiento_minimo

    def Formulacion(self):
        with open('Formaulación.txt', 'w') as archivo:
            archivo.write(str(self.modelo))

    # Generar Diagrama de Gantt
    def DiagramaGantt(self):
        '''
        DiagramaGantt()\n
        Genera un diagrama de Gantt para visualizar la programación resultante de las tareas y estaciones de trabajo.
        Notas:\n
        - Esta función genera un diagrama de Gantt para visualizar la programación resultante de las tareas y estaciones de trabajo obtenidas previamente del modelo de programación lineal (LP) resuelto.
        - Cada barra horizontal representa una tarea y su ubicación en el diagrama corresponde al tiempo de inicio de la tarea.
        - El eje vertical representa las estaciones de trabajo.
        - La longitud de cada barra representa la duración de la tarea.
        - Se muestra una leyenda para identificar cada tarea.\n
        Ejemplo:\n
            modelo.DiagramaGantt()
        '''
        self.ResumenResultados()
        fig, ax = plt.subplots(1)
        plt.title('Diagrama de Gantt')
        plt.xlabel('Tiempos de inicio')
        plt.ylabel('estaciones')
        for tareas in self.nombres_tareas:
            inicios = []
            estaciones = []
            duraciones = []
            for nombreInicio in self.tiempos_resultado.keys():
                if tareas in nombreInicio:
                    inicios.append(self.tiempos_resultado[nombreInicio])
                    tar, maq = nombreInicio.split('_',1)[0], nombreInicio.split('_',1)[1] 
                    duraciones.append(self.tareas[tar][maq])
                    estaciones.append(maq.split('_')[0])
            ax.barh(estaciones, duraciones, left=inicios, label=tareas)
        plt.legend(bbox_to_anchor=(1.02, 1.0), loc='upper left')
        plt.show()

class BalanceoLineaProgramacionLineal():
    def __init__(self, tareas, tiempoCiclo : int, objetivo = 'Numeros estaciones'):
        '''
        BalanceoLineaProgramacionLineal(tareas: dict[str, list[float, str]], tiempoCiclo: int, objetivo : str = 'Numeros estaciones')\n
        El objetivo es distribuir las tareas de manera equilibrada entre las estaciones de trabajo disponibles, con el fin de minimizar 
        los tiempos de espera y maximizar la eficiencia de la línea de producción.
        Argumentos:\n
        tareas : dict
            Enviar diccionario con nombre de cada tarea como llave, y cada valor debe ser una lista en la cual el primer valor debe ser
            un valor tipo int o float que representara el tiempo en SEGUNDOS de la tarea, y en la siguiente posicion de la lista un str
            con el nombre del predecesor, en caso de ser mas de un predecesor separarlos con el simbolo '-', es decir es un simbolo reservado 
            para su uso, y evitar el uso usar espacios
        tiempoCiclo: int
            Para el segundo argumento se debe enviar el timepo de ciclo, es el timepo necesario para completar una unidad de producción desde el incio hasta el final
        objetivo: str
            Función objetivo por la cual se resolvera el modelo, estan disponibles ['Numeros estaciones', ]
        Ejemplo:\n
            tareas={'A' : [12   , '-'],\n
                    'B' : [24   , '-'],\n
                    'C' : [42   , 'A'] ,\n
                    'D' : [6    , 'A-B'],\n
                    'E' : [18   , 'B'],\n
                    'F' : [6.6  , 'C'],\n
                    'G' : [19.2 , 'C'],\n
                    'H' : [36   , 'C-D'],\n
                    'I' : [16.2 , 'F-G-H'],\n
                    'J' : [22.8 , 'E-H'],\n
                    'K' : [30   , 'I-J'] }\n
            tiempo_ritmo =  60\n
            BalnceoLinea = BalanceoLineaProgramacionLineal(tareas, tiempo_ritmo, objetivo = 'Tiempo muerto')\n
            print(BalnceoLinea.DiccionarioEstaciones())\n
            print(BalnceoLinea.Metricas())\n
            BalnceoLinea.Formulacion()\n
            BalnceoLinea.Grafo(estaciones=True)
        '''
        self.funcion_objetivo = objetivo
        self.modelo = LpProblem("modelo_balanceo_linea", sense=LpMinimize)
        self.tareas = tareas
        self.tiempo_ritmo = tiempoCiclo
        self.EstacionesMinimas()
        self.CrearVariables()
        self.RestriccionPredecesores()
        self.RestriccionActivaciones()
        self.RestriccionTiempoCiclo()
        self.RestriccionTiempoMuerto()
        self.FuncionObjetivo()
        self.Solucionar()
    
    # Calcular minimo de estaciones sugeridas, mas un margen
    def EstacionesMinimas(self):
        if self.funcion_objetivo == 'Numeros estaciones':
            self.estaciones = sum([self.tareas[tarea][0] for tarea in self.tareas.keys()])/self.tiempo_ritmo
            self.estaciones = int((self.estaciones//1)+1)+1 # + 1 Extra para descartar infactibilidad del modelo
        else:
            self.estaciones = len(self.tareas)
    
    # Crear vairbales modelo
    def CrearVariables(self):
        self.binaria_estacion = LpVariable.dicts('BinariaEstacion', (estacion for estacion in range(self.estaciones)) , cat='Binary')
        self.binaria_tarea_estacion = LpVariable.dicts('BinariaTareaEstacion', ((tarea,estacion) for estacion in range(self.estaciones) for tarea in self.tareas.keys()) , cat='Binary')
        self.tiempo_ciclo = LpVariable('TiempoCiclo', lowBound=0, cat='Continuous')
        self.tiempo_ciclo_estacion = LpVariable.dicts('TiempoCicloEstacion', (estacion for estacion in range(self.estaciones)) , lowBound=0, cat='Continuous')
        self.tiempo_muerto_estacion = LpVariable.dicts('TiempoMuertoEstacion', (estacion for estacion in range(self.estaciones)) , lowBound=0, cat='Continuous')

    # Restriccion de predecesores
    def RestriccionPredecesores(self):
        for tarea in self.tareas.keys():
            if self.tareas[tarea][1] != '-':
                predecesores = self.tareas[tarea][1].split('-')
                for estacion in range(self.estaciones):
                    self.modelo += lpSum(self.binaria_tarea_estacion[predecesor,estacionacu] for predecesor in predecesores for estacionacu in range(0,estacion+1)) >= self.binaria_tarea_estacion[tarea,estacion]*len(predecesores) 

    # Restriccion Activaciones de estaciones y tareas por estacion
    def RestriccionActivaciones(self):
        for estacion in range(self.estaciones):
            self.modelo += lpSum(self.binaria_tarea_estacion[tarea,estacion]*self.tareas[tarea][0] for tarea in self.tareas.keys()) <= self.tiempo_ritmo*self.binaria_estacion[estacion]
            self.modelo += -(self.binaria_estacion[estacion]*1000000) + self.tiempo_ciclo_estacion[estacion] <= 0
        for tarea in self.tareas.keys():
            self.modelo += lpSum(self.binaria_tarea_estacion[tarea,estacion] for estacion in range(self.estaciones)) == 1

    # Calculo de tiempo de ciclo
    def RestriccionTiempoCiclo(self):
        for estacion in range(self.estaciones):
            self.modelo += lpSum(self.binaria_tarea_estacion[tarea,estacion]*self.tareas[tarea][0] for tarea in self.tareas.keys()) <= self.tiempo_ciclo 
            self.modelo += lpSum(self.binaria_tarea_estacion[tarea,estacion]*self.tareas[tarea][0] for tarea in self.tareas.keys()) == self.tiempo_ciclo_estacion[estacion] 

    # Calculo de tiempo muerto
    def RestriccionTiempoMuerto(self):
        for estacion in range(self.estaciones):
            self.modelo += self.tiempo_ciclo_estacion[estacion] + self.tiempo_muerto_estacion[estacion] >= self.tiempo_ciclo - (1-self.binaria_estacion[estacion])*1000000
    
    # Declaracion de Funcion objetivo
    def FuncionObjetivo(self):
        if self.funcion_objetivo == 'Numeros estaciones':
            self.modelo += lpSum(self.binaria_estacion[estacion] for estacion in range(self.estaciones))
        elif self.funcion_objetivo == 'Tiempo ciclo':
            self.modelo += self.tiempo_ciclo
        elif self.funcion_objetivo == 'Tiempo muerto':
            self.modelo += lpSum(self.tiempo_muerto_estacion[estacion] for estacion in range(self.estaciones))
        else:
            raise ValueError('Funcion objetivo no valida.')

    # Diccionario tareas por estacion
    def DiccionarioEstaciones(self):
        self.activacion_estacion = {}
        for v in self.modelo.variables():
            if 'BinariaTareaEstacion' in str(v) and v.varValue>0:
                nombre = str(v)
                nombre = nombre.replace('BinariaTareaEstacion_','').replace('(','').replace(')','').replace('_','').replace("'",'')
                nombre = nombre.split(',')
                self.activacion_estacion[nombre[0]] = 'Estacion '+ str(int(nombre[1])+1)
        return self.activacion_estacion

    def Metricas(self):
        self.metricas = pd.DataFrame()
        for v in self.modelo.variables():
            # BinariaEstacion
            if 'BinariaEstacion' in str(v):
                nombre = str(v)
                nombre = nombre.replace('BinariaEstacion_','').replace('(','').replace(')','').replace('_','').replace("'",'')
                nombre = nombre.split(',')
                if round(v.varValue,0) <= 0:
                    self.metricas.loc['Estacion '+ str(int(nombre[0])+1), 'Estación activa'] = 0
                else:
                    self.metricas.loc['Estacion '+ str(int(nombre[0])+1), 'Estación activa'] = int(v.varValue)
            # TiempoCicloEstacion 
            if 'TiempoCicloEstacion' in str(v):
                nombre = str(v)
                nombre = nombre.replace('TiempoCicloEstacion_','').replace('(','').replace(')','').replace('_','').replace("'",'')
                nombre = nombre.split(',')
                self.metricas.loc['Estacion '+ str(int(nombre[0])+1), 'Tiempo ciclo'] = round(v.varValue,1)
            # TiempoMuertoEstacion 
            if 'TiempoMuertoEstacion' in str(v):
                nombre = str(v)
                nombre = nombre.replace('TiempoMuertoEstacion_','').replace('(','').replace(')','').replace('_','').replace("'",'')
                nombre = nombre.split(',')
                self.metricas.loc['Estacion '+ str(int(nombre[0])+1), 'Tiempo muerto'] = round(v.varValue,1)
        self.metricas['Estación activa'] = self.metricas['Tiempo ciclo'].apply(lambda x: 0 if x==0 else 1)
        self.metricas['Tiempo muerto'] = self.metricas[['Tiempo ciclo','Tiempo muerto']].apply(lambda x: 0 if x['Tiempo ciclo']==0 else x['Tiempo muerto'], axis=1)
        # Eficiencia
        self.metricas['Eficiencia'] = self.metricas['Tiempo ciclo'] / self.metricas['Tiempo ciclo'].max()
        return self.metricas

    def Formulacion(self):
        with open('Formaulación.txt', 'w') as archivo:
            archivo.write(str(self.modelo))
    
    # Solucionar modelo multiobjetivo
    def Solucionar(self):
        if self.funcion_objetivo == 'Numeros estaciones':
            # Modelo Uso minimo de estaciones
            self.status = self.modelo.solve(solver=GUROBI(msg = False))
            if 'Optimal'== LpStatus[self.modelo.status]:
                self.horizonteTemporal = round(value(self.modelo.objective),0)
            else:
                raise 'Porblema en factibilidad del modelo'
            estaciones = 0
            # Asignacion de Restriccion Minima de Estaciones
            for v in self.modelo.variables():
                if 'BinariaEstacion' in str(v):
                    estaciones += v.varValue
            self.maximo_tiempo_estacion = LpVariable('MaximoTiempoEstacion', lowBound=0, cat='Continuous')
            for estacion in range(self.estaciones):
                self.modelo += lpSum(self.binaria_tarea_estacion[tarea,estacion]*self.tareas[tarea][0] for tarea in self.tareas.keys()) <= self.maximo_tiempo_estacion
            self.modelo += lpSum(self.binaria_estacion[estacion] for estacion in range(self.estaciones)) == estaciones 
            self.modelo += (1/sum([self.tareas[tarea][0] for tarea in self.tareas.keys()]))*(estaciones*self.maximo_tiempo_estacion)
        # Asignacion Maximizar Eficiencia de Linea en modelo
        self.status = self.modelo.solve(solver=GUROBI(msg = False))
        if 'Optimal'== LpStatus[self.modelo.status]:
            #print('-'*5+' Modelo solucionado correctamente '+'-'*5)
            self.horizonteTemporal = round(value(self.modelo.objective),0)
            self.DiccionarioEstaciones()
        else:
            raise 'Porblema en factibilidad del modelo'

    def Grafo(self, estaciones=False):
        # Crear un grafo dirigido
        G = nx.DiGraph()
        # Agregar nodos al grafo
        for tarea, datos in self.tareas.items():
            duracion, dependencias = datos
            G.add_node(tarea, duracion=duracion)
        # Agregar arcos entre las tareas dependientes
        for tarea, datos in self.tareas.items():
            dependencias = datos[1]
            if dependencias != '-':
                dependencias = dependencias.split('-')
                for dependencia in dependencias:
                    G.add_edge(dependencia, tarea)
        # Calcular el orden topológico
        orden_topologico = list(nx.topological_sort(G))
        nuevo_diccionario = {clave: valor[1].split('-') for clave, valor in self.tareas.items()}
        orden_topologico.reverse()
        # Crear un grafo dirigido con las dependencias
        Gaux = nx.DiGraph(nuevo_diccionario)
        # Crear un diccionario para agrupar los nodos por capa
        nodos_por_capa = {}
        for nodo in orden_topologico:
            predecesores = list(Gaux.predecessors(nodo))
            capa = max([nodos_por_capa[pre] for pre in predecesores], default=0) + 1
            nodos_por_capa[nodo] = capa
        # Organizar los nodos en listas por capa
        listas_por_capa = {}
        for nodo, capa in nodos_por_capa.items():
            if capa not in listas_por_capa:
                listas_por_capa[capa] = []
            listas_por_capa[capa].append(nodo)
        # Crear un nuevo gráfico ordenado con spring layout
        pos = {t: (c * -1500, p*-500) for c in listas_por_capa.keys() for p,t in enumerate(listas_por_capa[c])}
        # Dibujar el grafo ordenado
        if estaciones == True:
            # Crear diccionario de estaciones si no ha sido creado
            self.DiccionarioEstaciones()
            # Crear dicionario de colores
            colores_aleatorios = {nodo: "#{:02x}{:02x}{:02x}".format(np.random.randint(0, 255), np.random.randint(0, 255), np.random.randint(0, 255)) for nodo in set(self.activacion_estacion.values())}
            diccionario_colores = {t: colores_aleatorios[self.activacion_estacion[t]] for t in self.activacion_estacion.keys()}
            node_colors = [diccionario_colores[nodo] for nodo in G.nodes]
            nx.draw(G, pos=pos, with_labels=True, node_size=1000, node_color=node_colors, font_size=8, font_color='black')
        else:
            nx.draw(G, pos=pos, with_labels=True, node_size=1000, node_color='lightblue', font_size=8, font_color='black')
        # Mostrar el gráfico ordenado
        plt.show()

class SecuenciacionReglaJhonson():
    '''
    SecuenciacionReglaJhonson(tareas: dict[str ,dict[str:float, str:float, str:float]])\n
    Aplicacion de la regla de Jonson, para esto se debe enviar un diccioanrio de diccioanrios
    de la siguiente forma\n
    Argumentos:\n
    tareas: dict [dict]
        diccionario de key:tareas value:diccionarios de estaciones con duraciones 
        de cada tarea respectiva en la estacion, en este caso no todas las tareas deben ser procesadas en el mismo
        orden en a traves de cada una de las estaciones, esta heuristica acepta menor o igual a 3 estaciones\n
    Ejemplo:\n
        tareas={'T1' :{'M1':3,'M2':7,'M3':3},\n
                'T2' :{'M1':1,'M2':4,'M3':9},\n
                'T3' :{'M1':7,'M2':6,'M3':3},\n
                'T4' :{'M1':2,'M2':3,'M3':1},\n
                'T5' :{'M1':3,'M2':2,'M3':4},\n
                'T6' :{'M1':1,'M2':8,'M3':7},\n
                'T7' :{'M1':9,'M2':1,'M3':8},\n
                'T8' :{'M1':1,'M2':5,'M3':8},\n
                'T9' :{'M1':8,'M2':2,'M3':9},\n
                'T10':{'M1':6,'M2':1,'M3':7}}\n
        SecuneciaJhonson = SecuenciacionReglaJhonson(tareas)\n
        print(SecuneciaJhonson.combinaciones)\n
        print(SecuneciaJhonson.secuencias_posibles)\n
        print(SecuneciaJhonson.secuencias)\n
        print(SecuneciaJhonson.tiempos_procesos_secuencias)\n
        historico_secuencias = SecuneciaJhonson.Metricas().sort_values(['T. total fnalización'])
        print(historico_secuencias)\n
        SecuneciaJhonson.DiagramaGantt(historico_secuencias.iloc[0,0])\n
    '''
    def __init__(self, tareas):
        self.tareas_base = tareas
        self.tareas_base_original = tareas.copy()
        if len(tareas[list(tareas.keys())[0]])==2:
            self.tareas = tareas
        elif len(tareas[list(tareas.keys())[0]])==3:
            self.tareas={}
            estaciones = list(tareas[list(tareas.keys())[0]].keys())
            for tarea in tareas.keys():
                self.tareas[tarea] = {estaciones[0]+'-'+estaciones[1]:tareas[tarea][estaciones[0]]+tareas[tarea][estaciones[1]],
                estaciones[1]+'-'+estaciones[2]:tareas[tarea][estaciones[1]]+tareas[tarea][estaciones[2]]}
        else:
            raise 'El numero de tareas excede las 3 posibles que soluciona regla de Jhonson'
        self.nombres_tareas = list(self.tareas.keys())
        self.nombres_estaciones = list(list(self.tareas.values())[0].keys())
        self.EncontrarCombinaciones()
        self.CalcularPosibilidades()
        self.CalcularSecuencias()
        self.CalcularTiemposSecuencias()

    # Encontrar combinaciones posibles segun regla de Jhonson
    def EncontrarCombinaciones(self):
        self.combinaciones = []
        while self.tareas != {} :
            self.maximo = list(list(self.tareas.values())[0].values())[0]
            for tarea in self.tareas.keys():
                for estacion in self.nombres_estaciones:
                    if self.tareas[tarea][estacion] < self.maximo:
                        self.maximo = self.tareas[tarea][estacion]
            asignacion = []
            for tarea in self.tareas.keys():
                if self.tareas[tarea][self.nombres_estaciones[0]] == self.maximo:
                    asignacion.append([tarea,'I'])
                elif self.tareas[tarea][self.nombres_estaciones[1]] == self.maximo:
                    asignacion.append([tarea,'F'])
            tareas = list(set([tarea[0] for tarea in asignacion]))
            for tarea in tareas:
                self.tareas.pop(tarea)
            self.combinaciones.append(asignacion)
        return self.combinaciones

    # Calcular posibles combinaciones segun orden calculado
    def CalcularPosibilidades(self):
        self.secuencias_posibles = [[]]
        for combinacion in self.combinaciones:
            permutaciones = list(permutations(combinacion,len(combinacion)))
            for i in range(len(permutaciones)):
                permutaciones[i] = list(permutaciones[i])
            aux=[]
            for secuancia in self.secuencias_posibles:
                for posibilidad in permutaciones:
                    aux.append(secuancia+posibilidad)
            self.secuencias_posibles = aux
        return self.secuencias_posibles 

    # Calcular cada una de las seceuncias a partir de combinaciones de posibilidades
    def CalcularSecuencias(self):
        self.secuencias = []
        for secuencia in self.secuencias_posibles:
            inicio = []
            fin = []
            for tarea in secuencia:
                if tarea[1]=='F':
                    fin.insert(0, tarea[0])
                else:
                    inicio.append(tarea[0])
            self.secuencias.append(inicio+fin)
        return self.secuencias

    # Calcular tiempo de cada combinacion de posibilidad
    def CalcularTiemposSecuencias(self):
        self.tiempos_procesos_secuencias = []
        for secuencia in self.secuencias:
            self.tiempos_procesos_secuencias.append(self.CalcularTiempoProceso(secuencia))
        return self.tiempos_procesos_secuencias   

    # Calcular tiempo de proceso para cada secuencia
    def CalcularTiempoProceso(self, secuencia):
        duraciones = []
        for tarea in secuencia:
            duraciones.append([j for j in self.tareas_base_original[tarea].values()])
        matriz = [ [0 for j in i] for i in self.tareas_base_original.values()]
        for i in range(len(matriz)):
            for j in range(len(matriz[i])):
                if i==0 and j==0:
                    matriz[i][j] = duraciones[i][j]
                elif i==0:
                    matriz[i][j] = matriz[i][j-1] + duraciones[i][j]
                elif j==0:
                    matriz[i][j] = matriz[i-1][j] + duraciones[i][j]
                else:
                    matriz[i][j] = max([matriz[i][j-1],matriz[i-1][j]]) + duraciones[i][j]
        return matriz[i][j]
    
    # Grenerar df de tiempos inicio
    def TiemposInicio(self, secuencia):
        ti = pd.DataFrame(index= secuencia, columns=list(self.tareas_base_original[list(self.tareas_base_original.keys())[0]].keys()))
        for i, tar in enumerate(ti.index):
            for j, est in enumerate(ti.columns):
                if i == 0 and j == 0:
                    ti.loc[tar,est] = 0
                elif i == 0 and j != 0:
                    ti.loc[tar,est] = ti.iloc[i,j-1] + self.tareas_base_original[ti.index[i]][ti.columns[j-1]]
                elif i != 0 and j == 0:
                    ti.loc[tar,est] = ti.iloc[i-1,j] + self.tareas_base_original[ti.index[i-1]][ti.columns[j]]
                else:
                    ti.loc[tar,est] = max(ti.iloc[i-1,j] + self.tareas_base_original[ti.index[i-1]][ti.columns[j]], 
                                          ti.iloc[i,j-1] + self.tareas_base_original[ti.index[i]][ti.columns[j-1]])
        return ti
    
    # Metricas
    def Metricas(self):
        self.metricas_desempenio = pd.DataFrame()
        self.metricas_desempenio['Secuencia'] = self.secuencias
        self.metricas_desempenio['T. total fnalización'] = self.tiempos_procesos_secuencias
        for i,sec in enumerate(list(self.metricas_desempenio['Secuencia'])):
            ti = self.TiemposInicio(sec).values
            for e,est in enumerate(self.tareas_base_original[list(self.tareas_base_original.keys())[0]].keys()):
                #self.metricas_desempenio.loc[i,'T. inicio '+est]= ti[0][e]
                #self.metricas_desempenio.loc[i,'T. final '+est] = ti[-1][e] + self.tareas_base_original[sec[-1]][est]
                self.metricas_desempenio.loc[i,'T. bloqueo '+est] = ti[-1][e] + self.tareas_base_original[sec[-1]][est] - ti[0][e] - pd.DataFrame(self.tareas_base_original).T[est].sum()
        self.metricas_desempenio['T. total bloqueo'] = self.metricas_desempenio.iloc[:,2::].sum(axis=1)
        for i,sec in enumerate(list(self.metricas_desempenio['Secuencia'])):
            ti = self.TiemposInicio(sec).values
            for e,est in enumerate(self.tareas_base_original[list(self.tareas_base_original.keys())[0]].keys()):
                self.metricas_desempenio.loc[i, '%. utilización '+est] = (ti[-1][e] + self.tareas_base_original[sec[-1]][est] - ti[0][e]) / self.metricas_desempenio.loc[i,'T. total fnalización']
        return self.metricas_desempenio

    # Generar Diagrama de Gantt
    def DiagramaGantt(self, secuencia : list):
        '''
        DiagramaGantt()\n
        Genera un diagrama de Gantt para visualizar la programación resultante de las tareas y estaciones de trabajo.
        Notas:\n
        - Esta función genera un diagrama de Gantt para visualizar la programación resultante de las tareas y estaciones de trabajo obtenidas previamente del modelo de programación lineal (LP) resuelto.
        - Cada barra horizontal representa una tarea y su ubicación en el diagrama corresponde al tiempo de inicio de la tarea.
        - El eje vertical representa las estaciones de trabajo.
        - La longitud de cada barra representa la duración de la tarea.
        - Se muestra una leyenda para identificar cada tarea.\n
        Ejemplo:\n
            modelo.DiagramaGantt()
        '''
        ti = self.TiemposInicio(secuencia)
        self.tiempos_resultado = { tar+'_'+est : ti.loc[tar, est] for tar in ti.index for est in ti.columns}
        fig, ax = plt.subplots(1)
        plt.title('Diagrama de Gantt')
        plt.xlabel('Tiempos de inicio')
        plt.ylabel('estaciones')
        for tareas in self.tareas_base_original.keys():
            inicios = []
            estaciones = []
            duraciones = []
            for nombreInicio in self.tiempos_resultado.keys():
                if tareas in nombreInicio:
                    inicios.append(self.tiempos_resultado[nombreInicio])
                    tar, est = nombreInicio.split('_',1)[0], nombreInicio.split('_',1)[1] 
                    duraciones.append(self.tareas_base_original[tar][est])
                    estaciones.append(est.split('_')[0])
            ax.barh(estaciones, duraciones, left=inicios, label=tareas)
        plt.legend(bbox_to_anchor=(1.02, 1.0), loc='upper left')
        plt.show() 

class SecuenciacionReglaCDS():
    def __init__(self, tareas):
        '''
        SecuenciacionReglaCDS(tareas: dict[str ,dict[str:float, ..., str:float]])\n
        Aplicacion de la regla de CDS, para esto se debe enviar un diccioanrio de diccioanrios
        de la siguiente forma.\n
        Argumentos:\n
        tareas: dict [dict]
            diccionario de key:tareas value:diccionarios de estaciones con duraciones de cada tarea respectiva en la estacion, 
            en este caso no todas las tareas deben ser procesadas en el mismoorden en a traves de cada una de las estaciones, 
            esta heuristica acepta igual o mayor a 3 estaciones, solo que aplicara la misma regla de Jhonson.\n
        Ejemplo:\n
            tareas={'T1' :{'M1':3,'M2':7,'M3':3,'M4':3},\n
                    'T2' :{'M1':1,'M2':4,'M3':9,'M4':9},\n
                    'T3' :{'M1':7,'M2':6,'M3':3,'M4':1},\n
                    'T4' :{'M1':2,'M2':3,'M3':1,'M4':7},\n
                    'T5' :{'M1':3,'M2':2,'M3':4,'M4':2},\n
                    'T6' :{'M1':1,'M2':8,'M3':7,'M4':9},\n
                    'T7' :{'M1':9,'M2':1,'M3':8,'M4':4},\n
                    'T8' :{'M1':1,'M2':5,'M3':8,'M4':1},\n
                    'T9' :{'M1':8,'M2':2,'M3':9,'M4':4},\n
                    'T10':{'M1':6,'M2':1,'M3':7,'M4':4}}\n
            cds = SecuenciacionReglaCDS(tareas)\n
            print(cds.diccioanrio_tareas_orden)\n
            print(cds.combinaciones)\n
            print(cds.secuencias_posibles)\n
            print(cds.secuencias)\n
            print(cds.tiempos_procesos_secuencias)\n
            metricas_secuencias = cds.Metricas().sort_values(['T. total fnalización','T. total bloqueo'])\n
            print(metricas_secuencias)\n
            cds.DiagramaGantt(metricas_secuencias.iloc[0,0])\n
        '''
        self.tareas_base = tareas
        self.tareas_base_original = tareas.copy()
        self.tareas = tareas
        self.nombreTareas = list(tareas.keys())
        self.nombreestaciones = list(list(tareas.values())[0].keys())
        self.diccioanrio_tareas = []
        self.diccionario_tareas_orden = []
        self.combinaciones = []
        self.secuencias_posibles = []
        self.secuencias = []
        self.tiempos_procesos_secuencias = []
        for i in range(1,len(self.nombreestaciones)):
            estacion_ficticia_1 = self.nombreestaciones[0:i]
            estacion_ficticia_2 = self.nombreestaciones[-i::]
            tareasAuxiliar={}
            for tarea in self.nombreTareas:
                tareasAuxiliar[tarea] = {
                    '-'.join(estacion_ficticia_1):sum([self.tareas[tarea][estacion] for estacion in estacion_ficticia_1]),
                    '-'.join(estacion_ficticia_2):sum([self.tareas[tarea][estacion] for estacion in estacion_ficticia_2]),}
            self.diccioanrio_tareas.append(tareasAuxiliar)
            self.diccionario_tareas_orden.append(tareasAuxiliar.copy())
            combinaciones = self.EncontrarCombinaciones(tareasAuxiliar)
            self.combinaciones.append(combinaciones)
            posibilidades = self.CalcularPosibilidades(combinaciones)
            self.secuencias_posibles.append(posibilidades)
            secuencias = self.CalcularSecuencias(posibilidades) 
            self.secuencias.append(secuencias)
            tiempos = self.CalcularTiemposSecuencias(secuencias)
            self.tiempos_procesos_secuencias.append(tiempos)

    # Encontrar combinaciones posibles segun regla de Jhonson
    def EncontrarCombinaciones(self, tareasAuxiliar):
        Combinaciones = []
        while tareasAuxiliar!= {} :
            nombresestaciones = list(list(tareasAuxiliar.values())[0].keys())
            maximo = list(list(tareasAuxiliar.values())[0].values())[0]
            for tarea in tareasAuxiliar.keys():
                for estacion in nombresestaciones:
                    if tareasAuxiliar[tarea][estacion] < maximo:
                        maximo = tareasAuxiliar[tarea][estacion]
            asignacion = []
            for tarea in tareasAuxiliar.keys():
                if tareasAuxiliar[tarea][nombresestaciones[0]] == maximo:
                    asignacion.append([tarea,'I'])
                elif tareasAuxiliar[tarea][nombresestaciones[1]] == maximo:
                    asignacion.append([tarea,'F'])
            tareas = list(set([tarea[0] for tarea in asignacion]))
            for tarea in tareas:
                tareasAuxiliar.pop(tarea)
            Combinaciones.append(asignacion)
        return Combinaciones

    # Calcular posibles combinaciones segun orden calculado
    def CalcularPosibilidades(self, combinaciones):
        SecuenciasPosibles = [[]]
        for combinacion in combinaciones:
            permutaciones = list(permutations(combinacion,len(combinacion)))
            for i in range(len(permutaciones)):
                permutaciones[i] = list(permutaciones[i])
            auxiliar=[]
            for secuancia in SecuenciasPosibles:
                for posibilidad in permutaciones:
                    auxiliar.append(secuancia+posibilidad)
            SecuenciasPosibles = auxiliar
        return SecuenciasPosibles

    # Calcular cada una de las seceuncias a partir de combinaciones de posibilidades
    def CalcularSecuencias(self, SecuenciasPosibles):
        Secuencias = []
        for secuencia in SecuenciasPosibles:
            inicio = []
            fin = []
            for tarea in secuencia:
                if tarea[1]=='F':
                    fin.insert(0, tarea[0])
                else:
                    inicio.append(tarea[0])
            Secuencias.append(inicio+fin)
        return Secuencias
    
    # Calcular tiempo de cada combinacion de posibilidad
    def CalcularTiemposSecuencias(self, Secuencias):
        TiemposProcesosSecuencias = []
        for secuencia in Secuencias:
            TiemposProcesosSecuencias.append(self.CalcularTiempoProceso(secuencia))
        return TiemposProcesosSecuencias

    # Calcular tiempo de proceso para cada secuencia
    def CalcularTiempoProceso(self, secuencia):
        duraciones = []
        for tarea in secuencia:
            duraciones.append([j for j in self.tareas_base[tarea].values()])
        matriz = [ [0 for j in i] for i in self.tareas_base.values()]
        for i in range(len(matriz)):
            for j in range(len(matriz[i])):
                if i==0 and j==0:
                    matriz[i][j] = duraciones[i][j]
                elif i==0:
                    matriz[i][j] = matriz[i][j-1] + duraciones[i][j]
                elif j==0:
                    matriz[i][j] = matriz[i-1][j] + duraciones[i][j]
                else:
                    matriz[i][j] = max([matriz[i][j-1],matriz[i-1][j]]) + duraciones[i][j]
        return matriz[i][j]
    
    # Calcular tiempo de proceso para cada secuencia
    def CalcularTiempoProceso(self, secuencia):
        duraciones = []
        for tarea in secuencia:
            duraciones.append([j for j in self.tareas_base_original[tarea].values()])
        matriz = [ [0 for j in i] for i in self.tareas_base_original.values()]
        for i in range(len(matriz)):
            for j in range(len(matriz[i])):
                if i==0 and j==0:
                    matriz[i][j] = duraciones[i][j]
                elif i==0:
                    matriz[i][j] = matriz[i][j-1] + duraciones[i][j]
                elif j==0:
                    matriz[i][j] = matriz[i-1][j] + duraciones[i][j]
                else:
                    matriz[i][j] = max([matriz[i][j-1],matriz[i-1][j]]) + duraciones[i][j]
        return matriz[i][j]

    # Grenerar df de tiempos inicio
    def TiemposInicio(self, secuencia):
        ti = pd.DataFrame(index= secuencia, columns=list(self.tareas_base_original[list(self.tareas_base_original.keys())[0]].keys()))
        for i, tar in enumerate(ti.index):
            for j, est in enumerate(ti.columns):
                if i == 0 and j == 0:
                    ti.loc[tar,est] = 0
                elif i == 0 and j != 0:
                    ti.loc[tar,est] = ti.iloc[i,j-1] + self.tareas_base_original[ti.index[i]][ti.columns[j-1]]
                elif i != 0 and j == 0:
                    ti.loc[tar,est] = ti.iloc[i-1,j] + self.tareas_base_original[ti.index[i-1]][ti.columns[j]]
                else:
                    ti.loc[tar,est] = max(ti.iloc[i-1,j] + self.tareas_base_original[ti.index[i-1]][ti.columns[j]], ti.iloc[i,j-1] + self.tareas_base_original[ti.index[i]][ti.columns[j-1]])
        return ti
    
    # Metricas
    def Metricas(self):
        print('-'*100)
        self.metricas_desempenio = pd.DataFrame()
        self.metricas_desempenio['Secuencia'] = [sec for secuencias in self.secuencias for sec in secuencias]
        self.metricas_desempenio['T. total fnalización'] = [t_sec for timepos_secuencias in self.tiempos_procesos_secuencias for t_sec in timepos_secuencias]
        for i,sec in enumerate(list(self.metricas_desempenio['Secuencia'])):
            ti = self.TiemposInicio(sec).values
            for e,est in enumerate(self.tareas_base_original[list(self.tareas_base_original.keys())[0]].keys()):
                self.metricas_desempenio.loc[i,'T. bloqueo '+est] = ti[-1][e] + self.tareas_base_original[sec[-1]][est] - ti[0][e] - pd.DataFrame(self.tareas_base_original).T[est].sum()
        self.metricas_desempenio['T. total bloqueo'] = self.metricas_desempenio.iloc[:,2::].sum(axis=1)
        for i,sec in enumerate(list(self.metricas_desempenio['Secuencia'])):
            ti = self.TiemposInicio(sec).values
            for e,est in enumerate(self.tareas_base_original[list(self.tareas_base_original.keys())[0]].keys()):
                self.metricas_desempenio.loc[i, '%. utilización '+est] = (ti[-1][e] + self.tareas_base_original[sec[-1]][est] - ti[0][e]) / self.metricas_desempenio.loc[i,'T. total fnalización']
        return self.metricas_desempenio

    # Generar Diagrama de Gantt
    def DiagramaGantt(self, secuencia : list):
        '''
        DiagramaGantt()\n
        Genera un diagrama de Gantt para visualizar la programación resultante de las tareas y estaciones de trabajo.
        Notas:\n
        - Esta función genera un diagrama de Gantt para visualizar la programación resultante de las tareas y estaciones de trabajo obtenidas previamente del modelo de programación lineal (LP) resuelto.
        - Cada barra horizontal representa una tarea y su ubicación en el diagrama corresponde al tiempo de inicio de la tarea.
        - El eje vertical representa las estaciones de trabajo.
        - La longitud de cada barra representa la duración de la tarea.
        - Se muestra una leyenda para identificar cada tarea.\n
        Ejemplo:\n
            modelo.DiagramaGantt()
        '''
        ti = self.TiemposInicio(secuencia)
        self.tiempos_resultado = { tar+'_'+est : ti.loc[tar, est] for tar in ti.index for est in ti.columns}
        print(self.tiempos_resultado)
        fig, ax = plt.subplots(1)
        plt.title('Diagrama de Gantt')
        plt.xlabel('Tiempos de inicio')
        plt.ylabel('estaciones')
        for tareas in self.tareas_base_original.keys():
            inicios = []
            estaciones = []
            duraciones = []
            for nombreInicio in self.tiempos_resultado.keys():
                if tareas in nombreInicio:
                    inicios.append(self.tiempos_resultado[nombreInicio])
                    tar, est = nombreInicio.split('_',1)[0], nombreInicio.split('_',1)[1] 
                    duraciones.append(self.tareas_base_original[tar][est])
                    estaciones.append(est.split('_')[0])
            ax.barh(estaciones, duraciones, left=inicios, label=tareas)
        plt.legend(bbox_to_anchor=(1.02, 1.0), loc='upper left')
        plt.show()

class BranchAndBounds():
    def __init__(self, tareas : Dict):
        '''
        BranchAndBounds( tareas: dict, verbose : bool = False)
        Clase que implementa el algoritmo de Branch and Bound para resolver el problema de asignación de tareas en una línea de producción.\n
        Argumentos:\n
        tareas : dict
            Un diccionario que contiene las tareas y sus tiempos de procesamiento en diferentes estaciones de trabajo.
        Ejemplo:\n
            tareas={\n
            'T1' :{'M1':2,'M2':1,'M3':2},\n
            'T2' :{'M1':1,'M2':5,'M3':1},\n
            'T3' :{'M1':4,'M2':1,'M3':2},\n
            'T4' :{'M1':1,'M2':2,'M3':3},\n
            'T5' :{'M1':3,'M2':1,'M3':1},\n
            'T6' :{'M1':1,'M2':7,'M3':1},\n
            'T7' :{'M1':4,'M2':3,'M3':3},\n
            'T8' :{'M1':3,'M2':2,'M3':4}}\n
            bab = BranchAndBounds(tareas)\n
            print(bab.CalcularSecuencia(verbose = False))\n
            for key in bab.ecuaciones_bifurcaciones.keys():\n
                print(key)\n
                print(bab.ecuaciones_bifurcaciones[key])\n
            for key in bab.resultado_bifuraciones.keys():\n
                print(key)\n
                print(bab.resultado_bifuraciones[key])\n
            bab.CrearExcelResultado()\n
            bab.DiagramaGantt()\n
        '''
        self.tareas = tareas
        
        self.trabajos = list(self.tareas.keys())
        if len(self.trabajos) == 3 or len(self.trabajos) == 2:
            raise ValueError('Branch and Bouns solo permite cáluclo para 3 máquinas como máximo')
        self.procesos = {trabajo:list(self.tareas[trabajo].keys()) for trabajo in self.tareas.keys()}
        self.duraciones_original = pd.DataFrame(self.tareas).T
        self.duraciones = pd.DataFrame(self.tareas).T
        self.ecuaciones_bifurcaciones = {}
        self.resultado_bifuraciones = {}
        self.secuencia = []

    # Calcular arbol de decisión
    def CalcularSecuencia(self, verbose: bool = False):
        while len(self.duraciones.index)>1:
            self.CalcularRamificacionesDisponibles(verbose)
        self.secuencia.append(self.duraciones.index[0])
        return self.secuencia
    
    def CrearExcelResultado(self, ruta_excel : str = True):
        # Determinar ruta de creación de excel 
        if ruta_excel == True:
            self.ruta_excel_resultado = str(os.getcwd()).replace('\\','/')+'/Resultados Secuenciación Branch and Bound.xlsx'
        else:
            if ruta_excel[-5::] == '.xlsx':
                self.ruta_excel_resultado = ruta_excel
            else:
                raise 'Ruta no valida, no contiene la extension del archivo (<Nombre>.xlsx)'
        # Crear un nuevo libro de Excel
        libro = openpyxl.Workbook()
        hoja = libro.active
        # Variables para mantener el rastreo de la fila actual
        fila_actual = 2
        # Escribir formulación y resultados titulo
        celda = hoja.cell(row=fila_actual, column=2, value='FORMULACIÓN Y RESULTADOS BRANCH AND BOUND')
        celda.font = Font(bold=True, color="0000FF")
        # Calcular longitud
        ancho = len(self.resultado_bifuraciones[list(self.resultado_bifuraciones.keys())[0]].columns)
        # Inicializar contador de bifurcaciones
        bifurcacion = 0
        # Variables para mantener el rastreo de la fila actual
        fila_actual = 4
        for nombre_llave, df in self.resultado_bifuraciones.items():
            bifurcacion += 1
            # Escribir el nombre de la llave en la celda actual
            celda = hoja.cell(row=fila_actual, column=2, value=f'Bifurcacion {bifurcacion}')
            celda.font = Font(bold=True)
            celda.alignment  = Alignment(horizontal="center")
            fila_actual += 1
            # Escribir los nombres de las columnas
            for i, col_name in enumerate(df.columns, start=3):
                celda = hoja.cell(row=fila_actual, column=i, value=col_name)
                celda.font = Font(bold=True)
                celda.alignment = Alignment(horizontal="center")
            fila_actual += 1
            # Convertir el DataFrame en una tabla de Pandas y escribirlo en Excel
            for _, row in df.iterrows():
                celda = hoja.cell(row=fila_actual, column=2, value=row.name)
                celda.font = Font(bold=True)
                for i, value in enumerate(row, start=3):
                    celda = hoja.cell(row=fila_actual, column=i, value=value)
                    celda.alignment  = Alignment(horizontal="center")
                fila_actual += 1
            # Dejar una fila en blanco después de cada tabla
            fila_actual += 1
            # Escribir min max texto
            texto = str(df.values).replace('\n','').replace(' [',', max[').replace('[[','min[max[')
            celda = hoja.cell(row=fila_actual, column=2, value='min max')
            celda.font = Font(bold=True)
            celda.alignment = Alignment(horizontal="center")
            # Escribir min max valor
            celda = hoja.cell(row=fila_actual, column=3, value=df.apply(max, axis=1).min())
            celda.alignment = Alignment(horizontal="center")
            # Escribir min max ecuación
            celda = hoja.cell(row=fila_actual, column=2+ancho+2, value=texto)
            fila_actual += 1
            # Escribir secuenciar
            celda = hoja.cell(row=fila_actual, column=2, value='secuenciar')
            celda.font = Font(bold=True)
            celda.alignment = Alignment(horizontal="center")
            # Escribir trabajo a secuenciar
            celda = hoja.cell(row=fila_actual, column=3, value=nombre_llave)
            celda.alignment = Alignment(horizontal="center")
            # Dejar una fila en blanco después de cada tabla
            fila_actual += 2
        fila_actual = 4
        for nombre_llave, df in self.ecuaciones_bifurcaciones.items():
            fila_actual += 1
            # Escribir los nombres de las columnas
            for i, col_name in enumerate(df.columns, start=2+ancho+2):
                celda = hoja.cell(row=fila_actual, column=i, value=col_name)
                celda.font = Font(bold=True)
                celda.alignment  = Alignment(horizontal="center")
            fila_actual += 1
            # Convertir el DataFrame en una tabla de Pandas y escribirlo en Excel
            for _, row in df.iterrows():
                for i, value in enumerate(row, start=2+ancho+2):
                    celda = hoja.cell(row=fila_actual, column=i, value=value)
                    celda.alignment  = Alignment(horizontal="center")
                fila_actual += 1
            # Dejar una fila en blanco después de cada tabla
            fila_actual += 4
        # Guardar el libro de Excel
        libro.save(self.ruta_excel_resultado )
        libro.close()

    # Calular ramficacion Branch and Bound faltantes
    def CalcularRamificacionesDisponibles(self, verbose):
        df_ecuacion_bifurcacion = pd.DataFrame()
        df_resultado_bifurcacion = pd.DataFrame()
        # Suposición de cada trabajo como cuello de botella
        for trabajo in self.duraciones.index:
            # Mostrar los resultados por consola
            if verbose:
                print('-'*80)
            # Calculo de fechas de fin de trabajo LBj
            LBj = self.duraciones.cumsum(axis=1).T[[trabajo]].values
            P_t = self.duraciones.drop(trabajo)
            # Calculo de cotas de duración mínima
            for j,proceso in enumerate(P_t.columns):
                matriz_duraciones = P_t.values
                # Tiempos de trabajos no secuenciados en proceso cuello de botella
                duracion_proceso_cuello_botella = matriz_duraciones[:, j]
                # Tiempos de duración despues proceso considerado cuello de botella
                duracion_procesos_restantes = matriz_duraciones[:, j+1:]
                # Conversion de vectores en strings
                str_fin_trabajo_cuello_botella = LBj[j]
                str_tiempos_procesos_no_secuanciados_cuello_botella = str(duracion_proceso_cuello_botella).replace('\n',',').replace(' ','+')
                str_tiempo_despues_proceso_cuello_botella = str(duracion_procesos_restantes).replace('\n',',').replace(', ',',').replace(' ','+')
                str_ecuacion_LBjt = f'{LBj[j]} + sum{str_tiempos_procesos_no_secuanciados_cuello_botella} + min{str_tiempo_despues_proceso_cuello_botella}'
                # Mostrar los resultados por consola
                if verbose:
                    # Salida a texto
                    print(f'   Calculo cotas para trabajo {trabajo} y proceso cuello de botella {proceso}')
                    print('Fin trabajo proceso:', str_fin_trabajo_cuello_botella)
                    print('Tiempos de proceso no secuanciados:', str_tiempos_procesos_no_secuanciados_cuello_botella)
                    print('Duración mínima despues de cuello de botella:', str_tiempo_despues_proceso_cuello_botella)
                    print(f'LB-{trabajo}-{proceso}: {LBj[j]} + sum{str_tiempos_procesos_no_secuanciados_cuello_botella} + min{str_tiempo_despues_proceso_cuello_botella}')
                    print('')
                df_ecuacion_bifurcacion.loc[trabajo, proceso] = str_ecuacion_LBjt
                df_resultado_bifurcacion.loc[trabajo, proceso] = LBj[j][0] + np.sum(duracion_proceso_cuello_botella) + np.min(np.sum(duracion_procesos_restantes, axis = 1))
        trabajo_secuenciado = df_resultado_bifurcacion.max(axis=1).sort_values().head(1).index[0]
        self.secuencia.append(trabajo_secuenciado)
        # Mostrar los resultados por consola
        if verbose:
            print('-'*80)
            print('Resultado Calculos ecuaciones:\n', df_resultado_bifurcacion)
            print('Trabajo secuenciado:', trabajo_secuenciado)
        self.ecuaciones_bifurcaciones[trabajo_secuenciado] = df_ecuacion_bifurcacion
        self.resultado_bifuraciones[trabajo_secuenciado] = df_resultado_bifurcacion
        self.duraciones = self.duraciones.drop(trabajo_secuenciado, axis = 0)

    def CalcularDicciconarioTiempoSecuencia(self, secuencia):
        if secuencia == None:
            secuencia = self.secuencia
        diccionario_tiempos_inicio = {}
        df = self.duraciones_original.copy().reindex(secuencia)
        # Crear una matriz para almacenar los tiempos de inicio de cada tarea
        tiempo_inicio_tareas = pd.DataFrame(index=df.index, columns=df.columns)
        # Calcular el tiempo de inicio para la primera tarea en cada trabajo
        tiempo_inicio_tareas.iloc[0, 0] = df.iloc[0, 0]
        # Calcular el tiempo de inicio para el resto de las tareas en cada trabajo
        for i, trabajo in enumerate(df.index):
            for j, proceso in enumerate(df.columns):
                if i == 0 and j == 0:
                    tiempo_inicio_tareas.iloc[i, j] = 0
                if i == 0 and j != 0:
                    tiempo_inicio_tareas.iloc[i, j] = tiempo_inicio_tareas.iloc[i, j-1] + df.iloc[i, j-1]
                if i != 0 and j == 0:
                    tiempo_inicio_tareas.iloc[i, j] = tiempo_inicio_tareas.iloc[i-1, j] + df.iloc[i-1, j]
                if i != 0 and j != 0:
                    tiempo_inicio_tareas.iloc[i, j] = max(tiempo_inicio_tareas.iloc[i-1, j] + df.iloc[i-1, j], tiempo_inicio_tareas.iloc[i, j-1] + df.iloc[i, j-1])
                diccionario_tiempos_inicio[trabajo+'_'+proceso] = tiempo_inicio_tareas.iloc[i, j]
        return diccionario_tiempos_inicio

    # Generar Diagrama de Gantt
    def DiagramaGantt(self, secuencia: list = None):
        if secuencia == None:
            secuencia = self.secuencia
        diccionario_tiempos_inicio = self.CalcularDicciconarioTiempoSecuencia(secuencia)
        fig, ax = plt.subplots(1)
        plt.title('Diagrama de Gantt')
        plt.xlabel('Tiempos de inicio')
        plt.ylabel('estaciones')
        for trabajo in self.trabajos:
            inicios = []
            procesos = []
            duraciones = []
            for nombre_tiempo_inicio in diccionario_tiempos_inicio.keys():
                if trabajo in nombre_tiempo_inicio:
                    inicios.append(diccionario_tiempos_inicio[nombre_tiempo_inicio])
                    trabajo, proceso = nombre_tiempo_inicio.split('_',1)[0], nombre_tiempo_inicio.split('_',1)[1] 
                    duraciones.append(self.tareas[trabajo][proceso])
                    procesos.append(proceso.split('_')[0])
            ax.barh(procesos, duraciones, left=inicios, label=trabajo)
        plt.legend(bbox_to_anchor=(1.02, 1.0), loc='upper left')
        plt.show()

class ReglasSecuenciacion():
    def __init__(self, tareas : Dict, importancia : pd.DataFrame = None):
        '''
        Clase para aplicar diferentes reglas de secuenciación a un conjunto de tareas.\n
        Argumentos:\n
        tareas : dict
            Diccionario que contiene las tareas a ser secuenciadas. Cada tarea se representa mediante un diccionario que incluye: duración, tiempo de inicio, tiempo de entrega y costos de retraso.\n
        importancia : DataFrame
            DataFrame que contiene el indice como cada tarea y una unica columnas con las importancias de cada tarea qy esta debe sumar 1.\n
        Ejmplo:\n
            tareas={'T1' : {'duraciones':1, 'timpos inicio':4  , 'tiempo entrega':8,  'costos retraso':1},\n
                    'T2' : {'duraciones':5, 'timpos inicio':1  , 'tiempo entrega':12, 'costos retraso':8},\n
                    'T3' : {'duraciones':1, 'timpos inicio':2  , 'tiempo entrega':29, 'costos retraso':7},\n
                    'T4' : {'duraciones':2, 'timpos inicio':3  , 'tiempo entrega':15, 'costos retraso':5},\n
                    'T5' : {'duraciones':1, 'timpos inicio':5  , 'tiempo entrega':28, 'costos retraso':3},\n
                    'T6' : {'duraciones':7, 'timpos inicio':10 , 'tiempo entrega':19, 'costos retraso':1},\n
                    'T7' : {'duraciones':3, 'timpos inicio':6  , 'tiempo entrega':15, 'costos retraso':9},\n
                    'T8' : {'duraciones':2, 'timpos inicio':8  , 'tiempo entrega':9,  'costos retraso':4}}\n
            importancias = pd.DataFrame([0.05,0.30,0.15,0.1,0.1,0.1,0.1,0.1], index=list(tareas.keys()))\n
            rs = ReglasSecuenciacion(tareas, importancias)\n
            print(rs.TiempoProcesamientoCorto())\n
            print(rs.TiempoProcesamientoLargo())\n
            print(rs.TiempoFinalizacionTemprano())\n
            print(rs.TiempoFinalizacionTardio())\n
            print(rs.RelacionProcesameintoEntregaCorto())\n
            print(rs.RelacionProcesameintoEntregaLargo())\n
            print(rs.CostoRestrasoAlto())\n
            for nom, val in rs.ReglasSecuenciacion().items():\n
                print(nom, '\n', val)\n
            print(rs.MetricasGeneralesSecuenciacion())\n
            print(rs.GraficarMetricas())\n
        '''
        self.tareas = tareas
        self.data = pd.DataFrame(tareas).T
        self.importancia = pd.DataFrame(index = self.data.index)
        if importancia is None:
            self.importancia['importancia'] = 1/len(self.data.index)
        else:
            # Ordenar los índices de ambos DataFrames
            dfaux1 = self.data.sort_index()
            dfaux2 = importancia.sort_index()
            if dfaux1.index.equals(dfaux2.index):
                if importancia[importancia.columns[0]].sum() != 1:
                    raise("Las sumatoria de las importancias no es igual a 1")
            else:
                raise("Los índices de los DataFrames son diferentes")
            self.importancia['importancia'] = importancia

    # Calcular metricas
    def Metricas(self, df : pd.DataFrame):
        df['inicio real'] = 0
        df['final real'] = 0
        tareas = list(df.index)
        for i, ind in enumerate(df.index):
            if i == 0:
                df.loc[ind,'inicio real'] = df.iloc[i,1]
                df.loc[ind,'final real'] = df.loc[ind,'inicio real'] + df.iloc[i,0]
            else:
                df.loc[ind,'inicio real'] = max(df.iloc[i,1], df.loc[tareas[i-1],'final real'])
                df.loc[ind,'final real'] = df.loc[ind,'inicio real'] + df.iloc[i,0]
        df['tiempo flujo'] = df['final real'] * df['importancia']
        for i, ind in enumerate(df.index):
            df.loc[ind,'anticipación'] = (df.iloc[i,2]-df.loc[ind,'final real'])*df.loc[ind,'importancia'] if df.iloc[i,2]-df.loc[ind,'final real']>0 else 0
            df.loc[ind,'tardanza'] = (df.loc[ind,'final real']-df.iloc[i,2])*df.loc[ind,'importancia'] if df.loc[ind,'final real']-df.iloc[i,2]>0 else 0
        return df
    
    #LIFO
    #FIFO

    # SPT: Tiempo de Procesamiento más Corto
    def TiempoProcesamientoCorto(self):
        self.tiempo_procesamiento_corto = self.data.copy().sort_values([self.data.columns[0]], ascending=True)
        self.tiempo_procesamiento_corto['importancia'] = self.importancia
        self.tiempo_procesamiento_corto = self.Metricas(self.tiempo_procesamiento_corto)
        return self.tiempo_procesamiento_corto
    
    # LPT: Tiempo de Procesamiento más Largo
    def TiempoProcesamientoLargo(self):
        self.tiempo_procesamiento_largo = self.data.copy().sort_values([self.data.columns[0]], ascending=False)
        self.tiempo_procesamiento_largo['importancia'] = self.importancia
        self.tiempo_procesamiento_largo = self.Metricas(self.tiempo_procesamiento_largo)
        return self.tiempo_procesamiento_largo

    # EFT: Tiempo de Finalización más Temprano
    def TiempoFinalizacionTemprano(self):
        self.tiempo_finalizacion_temprano = self.data.copy()
        self.tiempo_finalizacion_temprano['tiempo finalizacion'] = self.tiempo_finalizacion_temprano[self.tiempo_finalizacion_temprano.columns[1]] + self.tiempo_finalizacion_temprano[self.tiempo_finalizacion_temprano.columns[0]]
        self.tiempo_finalizacion_temprano = self.tiempo_finalizacion_temprano.sort_values([self.tiempo_finalizacion_temprano.columns[4]], ascending=True)
        self.tiempo_finalizacion_temprano['importancia'] = self.importancia
        self.tiempo_finalizacion_temprano = self.Metricas(self.tiempo_finalizacion_temprano)
        return self.tiempo_finalizacion_temprano
    
    # LFT: Tiempo de Finalización más Tardío
    def TiempoFinalizacionTardio(self):
        self.tiempo_finalizacion_tardio = self.data.copy()
        self.tiempo_finalizacion_tardio['tiempo finalizacion'] = self.tiempo_finalizacion_tardio[self.tiempo_finalizacion_tardio.columns[1]] + self.tiempo_finalizacion_tardio[self.tiempo_finalizacion_tardio.columns[0]]
        self.tiempo_finalizacion_tardio = self.tiempo_finalizacion_tardio.sort_values([self.tiempo_finalizacion_tardio.columns[4]], ascending=True)
        self.tiempo_finalizacion_tardio['importancia'] = self.importancia
        self.tiempo_finalizacion_tardio = self.Metricas(self.tiempo_finalizacion_tardio)
        return self.tiempo_finalizacion_tardio
    
    # EDD: Tiempo de Entrega (Earliest Due Date)
    def ProcesameintoEntregaCorto(self):
        self.procesamiento_entrega_corto = self.data.copy()
        self.procesamiento_entrega_corto = self.procesamiento_entrega_corto.sort_values([self.procesamiento_entrega_corto.columns[2]], ascending=True)
        self.procesamiento_entrega_corto['importancia'] = self.importancia
        self.procesamiento_entrega_corto = self.Metricas(self.procesamiento_entrega_corto)
        return self.procesamiento_entrega_corto

    # S/EDD: Relación Tiempo de Procesamiento-Tiempo de Entrega (Corto/Earliest Due Date)
    def RelacionProcesameintoEntregaCorto(self):
        self.relacion_procesamiento_entrega_corto = self.data.copy()
        self.relacion_procesamiento_entrega_corto['relacion procesamiento entrega'] = self.relacion_procesamiento_entrega_corto[self.relacion_procesamiento_entrega_corto.columns[0]]/self.relacion_procesamiento_entrega_corto[self.relacion_procesamiento_entrega_corto.columns[2]]
        self.relacion_procesamiento_entrega_corto = self.relacion_procesamiento_entrega_corto.sort_values([self.relacion_procesamiento_entrega_corto.columns[4]], ascending=True)
        self.relacion_procesamiento_entrega_corto['importancia'] = self.importancia
        self.relacion_procesamiento_entrega_corto = self.Metricas(self.relacion_procesamiento_entrega_corto)
        return self.relacion_procesamiento_entrega_corto
    
    # L/EDD: Relación Tiempo de Procesamiento-Tiempo de Entrega (Largo/Earliest Due Date)
    def RelacionProcesameintoEntregaLargo(self):
        self.relacion_procesamiento_entrega_largo = self.data.copy()
        self.relacion_procesamiento_entrega_largo['relacion procesamiento entrega'] = self.relacion_procesamiento_entrega_largo[self.relacion_procesamiento_entrega_largo.columns[0]]/self.relacion_procesamiento_entrega_largo[self.relacion_procesamiento_entrega_largo.columns[2]]
        self.relacion_procesamiento_entrega_largo = self.relacion_procesamiento_entrega_largo.sort_values([self.relacion_procesamiento_entrega_largo.columns[4]], ascending=False)
        self.relacion_procesamiento_entrega_largo['importancia'] = self.importancia
        self.relacion_procesamiento_entrega_largo = self.Metricas(self.relacion_procesamiento_entrega_largo)
        return self.relacion_procesamiento_entrega_largo

    # HLC: Costo de Retraso más Alto
    def CostoRestrasoAlto(self):
        self.costo_retraso_alto = self.data.copy().sort_values([self.data.columns[3]], ascending=False)
        self.costo_retraso_alto['importancia'] = self.importancia
        self.costo_retraso_alto = self.Metricas(self.costo_retraso_alto)
        return self.costo_retraso_alto
    
    # Todos las reglas de secuanciacion
    def ReglasSecuenciacion(self):
        self.diccionario_reglas_secuanciacion = {}
        try:
            self.diccionario_reglas_secuanciacion['SPT tiempo procesamiento mas corto'] = self.TiempoProcesamientoCorto()
        except:
            pass
        try:
            self.diccionario_reglas_secuanciacion['LPT tiempo procesamiento mas largo'] = self.TiempoProcesamientoLargo()
        except:
            pass
        try:
            self.diccionario_reglas_secuanciacion['EFT tiempo finalizacion mas temprano'] = self.TiempoFinalizacionTemprano()
        except:
            pass
        try:
            self.diccionario_reglas_secuanciacion['LFT tiempo finalizacion mas tardio'] = self.TiempoFinalizacionTardio()
        except:
            pass
        try:
            self.diccionario_reglas_secuanciacion['EDD tiempo entrega mas corto'] = self.ProcesameintoEntregaCorto()
        except:
            pass
        try:
            self.diccionario_reglas_secuanciacion['S/EDD relacion tiempo procesamiento con entrega mas corto'] = self.RelacionProcesameintoEntregaCorto()
        except:
            pass
        try:
            self.diccionario_reglas_secuanciacion['L/EDD relacion tiempo procesamiento con entrega mas largo'] = self.RelacionProcesameintoEntregaLargo()
        except:
            pass
        try:
            self.diccionario_reglas_secuanciacion['HLC costo retraso mas alto'] = self.CostoRestrasoAlto()
        except:
            pass
        return self.diccionario_reglas_secuanciacion
    
    # Todos las metricas reglas de secuanciacion
    def MetricasGeneralesSecuenciacion(self):
        self.metricas_generales = pd.DataFrame()
        try:
            self.metricas_generales.loc['SPT tiempo procesamiento mas corto','tiempo flujo'] = self.TiempoProcesamientoCorto()['tiempo flujo'].sum()
            self.metricas_generales.loc['SPT tiempo procesamiento mas corto','anticipación'] = self.TiempoProcesamientoCorto()['anticipación'].sum()
            self.metricas_generales.loc['SPT tiempo procesamiento mas corto','tardanza'    ] = self.TiempoProcesamientoCorto()['tardanza'].sum()    
        except:
            pass
        try:
            self.metricas_generales.loc['LPT tiempo procesamiento mas largo','tiempo flujo'] = self.TiempoProcesamientoLargo()['tiempo flujo'].sum()
            self.metricas_generales.loc['LPT tiempo procesamiento mas largo','anticipación'] = self.TiempoProcesamientoLargo()['anticipación'].sum()
            self.metricas_generales.loc['LPT tiempo procesamiento mas largo','tardanza'    ] = self.TiempoProcesamientoLargo()['tardanza'].sum()    
        except:
            pass
        try:
            self.metricas_generales.loc['EFT tiempo finalizacion mas temprano','tiempo flujo'] = self.TiempoFinalizacionTemprano()['tiempo flujo'].sum()
            self.metricas_generales.loc['EFT tiempo finalizacion mas temprano','anticipación'] = self.TiempoFinalizacionTemprano()['anticipación'].sum()
            self.metricas_generales.loc['EFT tiempo finalizacion mas temprano','tardanza'    ] = self.TiempoFinalizacionTemprano()['tardanza'].sum()    
        except:
            pass
        try:
            self.metricas_generales.loc['LFT tiempo finalizacion mas tardio','tiempo flujo'] = self.TiempoFinalizacionTardio()['tiempo flujo'].sum()
            self.metricas_generales.loc['LFT tiempo finalizacion mas tardio','anticipación'] = self.TiempoFinalizacionTardio()['anticipación'].sum()
            self.metricas_generales.loc['LFT tiempo finalizacion mas tardio','tardanza'    ] = self.TiempoFinalizacionTardio()['tardanza'].sum()    
        except:
            pass
        try:
            self.metricas_generales.loc['EDD tiempo entrega mas corto','tiempo flujo'] = self.ProcesameintoEntregaCorto()['tiempo flujo'].sum()
            self.metricas_generales.loc['EDD tiempo entrega mas corto','anticipación'] = self.ProcesameintoEntregaCorto()['anticipación'].sum()
            self.metricas_generales.loc['EDD tiempo entrega mas corto','tardanza'    ] = self.ProcesameintoEntregaCorto()['tardanza'].sum()    
        except:
            pass
        try:
            self.metricas_generales.loc['S/EDD relacion tiempo procesamiento con entrega mas corto','tiempo flujo'] = self.RelacionProcesameintoEntregaCorto()['tiempo flujo'].sum()
            self.metricas_generales.loc['S/EDD relacion tiempo procesamiento con entrega mas corto','anticipación'] = self.RelacionProcesameintoEntregaCorto()['anticipación'].sum()
            self.metricas_generales.loc['S/EDD relacion tiempo procesamiento con entrega mas corto','tardanza'    ] = self.RelacionProcesameintoEntregaCorto()['tardanza'].sum()    
        except:
            pass
        try:
            self.metricas_generales.loc['L/EDD relacion tiempo procesamiento con entrega mas largo','tiempo flujo'] = self.RelacionProcesameintoEntregaLargo()['tiempo flujo'].sum()
            self.metricas_generales.loc['L/EDD relacion tiempo procesamiento con entrega mas largo','anticipación'] = self.RelacionProcesameintoEntregaLargo()['anticipación'].sum()
            self.metricas_generales.loc['L/EDD relacion tiempo procesamiento con entrega mas largo','tardanza'    ] = self.RelacionProcesameintoEntregaLargo()['tardanza'].sum()    
        except:
            pass
        try:
            self.metricas_generales.loc['HLC costo retraso mas alto','tiempo flujo'] = self.CostoRestrasoAlto()['tiempo flujo'].sum()
            self.metricas_generales.loc['HLC costo retraso mas alto','anticipación'] = self.CostoRestrasoAlto()['anticipación'].sum()
            self.metricas_generales.loc['HLC costo retraso mas alto','tardanza'    ] = self.CostoRestrasoAlto()['tardanza'].sum()    
        except:
            pass
        return self.metricas_generales
    
    # Grafico de metricas
    def GraficarMetricas(self):
        df = self.MetricasGeneralesSecuenciacion()
        # Definir las categorías
        categorias = [x.split(' ')[0] for x in df.index]
        # Crear una lista de ángulos para el gráfico de telaraña
        angulos = [n / float(len(categorias)) * 2 * np.pi for n in range(len(categorias))]
        angulos += angulos[:1]
        # Crear subplots, uno para cada columna
        fig, axs = plt.subplots(1, 3, figsize=(12, 4), subplot_kw={'polar': True})  # 1 fila y 3 columnas de subplots
        # Iterar a través de las columnas y crear un gráfico de telaraña para cada una
        colores = ['#FFFF00','#008000','#FF0000']
        for i, columna in enumerate(df.columns):  # Ignorar la columna 'Categoria'
            puntuaciones = df[columna].tolist()
            puntuaciones += puntuaciones[:1]
            axs[i].fill(angulos, puntuaciones, alpha=0.25, color=colores[i])
            axs[i].set_xticks(angulos[:-1])
            axs[i].set_xticklabels(categorias)
            axs[i].set_title(columna)
        # Ajustar el espaciado entre los gráficos
        plt.tight_layout()
        # Mostrar los gráficos
        plt.show()