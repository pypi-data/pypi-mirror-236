import argparse
import csv
import re

from openpyxl import load_workbook

parser = argparse.ArgumentParser(
    description='Converts the first sheet of an .xlsx-file to a .csv-file'
)
parser.add_argument('xlsxfile', help='The .xlsx file to be converted')
parser.add_argument('-o,--output', dest='output', help='Name of the output file')
parser.add_argument(
    '-s,--sheet',
    dest='sheet',
    help='Name of sheet to convert to .xlsx. By default the first sheet will be converted',
)


def main():
    args = parser.parse_args()

    csv_rows = []

    wb = load_workbook(args.xlsxfile)

    sheet_index = 0
    if args.sheet:
        assert args.sheet in wb.sheetnames
        sheet_index = wb.sheetnames.index(args.sheet)

    ws = None
    for sheet in wb:
        if args.sheet:
            if sheet.title == args.sheet:
                ws = sheet
                break
        else:
            ws = sheet

    if ws is None:
        print(f'Sheet with name "{args.name}" not found')
        exit(1)

    for xls_row in ws:
        csv_row = []
        for xls_cell in xls_row:
            csv_row.append(xls_cell.value)
        csv_rows.append(csv_row)

    if args.output:
        output_filename = args.output
    else:
        output_filename = re.sub(r'\.xlsx$', '.csv', args.xlsxfile)
    with open(output_filename, 'w') as output_file:
        writer = csv.writer(output_file)
        writer.writerows(csv_rows)
