#!/usr/bin/env python
# -*- coding: utf-8 -*-
########################################################################
#
# Copyright (c) 2023 datavita.com.cn, Inc. All Rights Reserved
#
########################################################################


import pandas as pd
import numpy as np
import openpyxl
import traceback
from ..constant import UNIV_NAME, RET_TYPE, IND_DICT
import fqdatac
import duckdb
import os
from dataclasses import dataclass
from functools import reduce
from datetime import datetime

from fqdatac.etl.db_tools import timer


class FQDataLoader(object):
    def __init__(self):
        self.today = datetime.now().strftime("%Y%m%d")
        pd.set_option("display.max_columns", 100)
        pd.set_option("display.width", 1000)
        pd.set_option("display.max_colwidth", 100)
        pd.set_option("display.float_format", lambda x: "%.4f" % x)
        addr_prod = ("api.datavita.com.cn", 443)
        addr_dev = ("219.143.244.230", 8390)
        fqdatac.init(access_key="123", access_secret="456", addr=addr_dev, timeout=60)

    @staticmethod
    def partition_df(df: pd.DataFrame, col: str):
        """Partition a DataFrame into a dictionary of DataFrames by column value."""
        return {k: v.reset_index(drop=True) for k, v in df.groupby(col, sort=False)}


@dataclass
class Factor:
    """
    一个因子的全部数据类
    """

    hist: pd.DataFrame

    def __post_init__(self):
        assert set(["K", "D", "T"]) < set(self.hist.columns)
        self.facname = list(set(self.hist.columns) - set(["K", "D", "T"]))[0]
        self.hist = self.hist.sort_values(by=["D", "T"])
        self.hist_d = FQDataLoader.partition_df(self.hist, "D")
        self.facdays = sorted(list(set(self.hist["D"])))

    @property
    def coverage(self):
        return self.hist["K"].nunique()

    @property
    def sd(self):
        return self.facdays[0]

    @property
    def ed(self):
        return self.facdays[-1]

    @property
    def numdays(self):
        return len(fqdatac.get_daysrange(self.sd, self.ed))

    @property
    def hist_k(self):
        return FQDataLoader.partition_df(self.hist, "K")

    @classmethod
    def read_factor(cls, path, sd):
        factor_data = []
        for file in os.listdir(path):
            dt, fmt = file.split(".")
            if dt >= sd and fmt == "csv":
                factor_data.append(pd.read_csv(os.path.join(path, file)))
        alldata = pd.concat(factor_data, axis=0)
        return cls(hist=alldata)

    @classmethod
    def from_dict(cls, d: dict):
        factor_data = [v for v in d.values()]
        alldata = pd.concat(factor_data, axis=0)
        return cls(hist=alldata)

    def save_tocsv(self, path):
        os.makedirs(path, exist_ok=True)
        for k, v in self.hist_d.items():
            v.to_csv(os.path.join(path, f"{k}.csv"))


@dataclass
class FactorFile:
    """
    一个因子某单日的因子文件
    """

    kdtv: pd.DataFrame
    fac_name: str
    dt: str

    def __post_init__(self):
        assert set(["K", "D", "T"]) < set(self.kdtv.columns)


class FactorBacktest(FQDataLoader):
    def __init__(
            self, sd, ed, backtest_name, univ_name, fdays, rettype, directory_list
    ):
        self.backtest_day = datetime.now().strftime("%Y%m%d")
        self.sd = sd
        self.ed = ed
        self.backtest_name = backtest_name
        self.univ_name = univ_name
        self.fdays = fdays
        self.rettype = rettype
        self.directory_list = directory_list
        self.output_dir = "/data/quant/report/factor_backtest"

    def check_backtest_legitimacy(self):
        """
        check to see if legitimate backtest inputs are given
        """
        # univ_name必须在预设的list里面
        try:
            assert self.univ_name in UNIV_NAME
        except AssertionError:
            raise ValueError(f"preset univ_name {self.univ_name} not found")

        # ret_type必须在预设的list里面
        try:
            assert self.rettype in RET_TYPE
        except AssertionError:
            raise ValueError(f"preset ret_type {self.rettype} not found")

        # 向前预测期必须在1~63天的整数之内
        try:
            assert isinstance(self.fdays, int) and 1 <= self.fdays <= 63
        except AssertionError:
            raise ValueError("forward looking days should fall into range 1~63 days")

        # 必须至少传一个存在的factor路径
        try:
            assert len(self.directory_list) >= 1
        except AssertionError:
            raise ValueError("must pass at least 1 valid factor path")

    def checkdates(self):
        """
        查回测日期区间
        """
        try:
            assert isinstance(self.ed, str) and isinstance(self.sd, str)
            assert fqdatac.get_day(self.ed, -60) >= self.sd
            assert self.ed <= fqdatac.get_day(self.backtest_day, -1)
            assert self.sd >= "20130000"
        except AssertionError:
            raise KeyError("invalid date range inputs")

    def get_univ_cnall(self):
        univ = fqdatac.get_universe(sd=self.sd, ed=self.ed, univ_name="cnall")
        return self.partition_df(univ, "D")

    def get_univ_userselect(self):
        univ = fqdatac.get_universe(sd=self.sd, ed=self.ed, univ_name=self.univ_name)
        return self.partition_df(univ, "D")


class FactorBTLoader(FactorBacktest):

    """
    因子回测时的加载类
    """

    def __init__(
            self, sd, ed, backtest_name, univ_name, fdays, rettype, directory_list
    ):
        super(FactorBTLoader, self).__init__(
            sd, ed, backtest_name, univ_name, fdays, rettype, directory_list
        )
        self.check_backtest_legitimacy()
        self.checkdates()
        self.container = dict()
        self.alldays = fqdatac.get_daysrange(self.sd, self.ed)
        self.cnall = self.get_univ_cnall()
        self.univ = self.get_univ_userselect()

    def load(self, directory) -> dict:
        """
        将因子load进container
        """
        self.checker = FactorFileChecker(
            sd=self.sd,
            ed=self.ed,
            backtest_name=self.backtest_name,
            univ_name=self.univ_name,
            fdays=self.fdays,
            rettype=self.rettype,
            directory_list=self.directory_list,
            directory=directory,
        )

        files_to_load = self.checker.checkfiles()  # 查路径里的文件是否齐全 是否重复
        fullpath = [os.path.join(directory, x) for x in files_to_load]
        facdict = {}
        for f in fullpath:
            dtcsv = pd.read_csv(f)
            dtcsv["D"] = dtcsv["D"].astype(str)
            self.checker.checkformat(f, dtcsv)  # 查一下文件结构是否符合规范
            facdict[dtcsv["D"].iloc[0]] = dtcsv  # 确保一个文件里只有一个D，已经查过了
        self.container[self.checker.facname] = facdict

    def onlyuniv(self):
        """
        只保留当天在universe中的股票进行测试
        """
        for k, v in self.container.items():
            for dt in list(v.keys()):
                v[dt] = v[dt][v[dt]["K"].isin(list(self.univ[dt]["K"]))]

    def fwdret_load(self):
        """
        获取对照的股票return数据
        """
        self.fwdret1 = None
        self.fwdret = None
        self.fwdret = fqdatac.get_bfwd(
            sd=self.sd,
            ed=self.ed,
            univ_name=self.univ_name,
            ret_type=self.rettype,
            dayshift=self.fdays,
        )
        # self.fwdret = self.partition_df(fwdret, "D")
        if self.fdays == 1:
            self.fwdret1 = self.fwdret.copy()
        else:
            self.fwdret1 = fqdatac.get_bfwd(
                sd=self.sd,
                ed=self.ed,
                univ_name=self.univ_name,
                ret_type=self.rettype,
                dayshift=1,
            )
        #     self.fwdret1 = self.partition_df(fwdret1, "D")


class FactorFileChecker(FactorBTLoader):

    """
    因子路径、文件的格式校验
    """

    def __init__(
            self,
            sd,
            ed,
            backtest_name,
            univ_name,
            fdays,
            rettype,
            directory_list,
            directory,
    ):
        super(FactorFileChecker, self).__init__(
            sd, ed, backtest_name, univ_name, fdays, rettype, directory_list
        )
        self.path = directory
        self.facname = self.path.split("/")[-1]

    def checkfiles(self):
        """
        查重复文件，查缺失文件
        """
        files = os.listdir(self.path)
        csv_files = sorted([x for x in files if x.endswith(".csv")])

        files_alldays = [f"{x}.csv" for x in self.alldays]
        files_shouldbe = sorted([x for x in files_alldays if x >= csv_files[0]])
        files_existing = sorted([x for x in csv_files if x in files_alldays])

        # 查找重复文件
        counts = [files_existing.count(item) for item in files_existing]
        if max(counts) >= 2:
            raise FileExistsError("duplicate files of the same day")

        # 查找缺失文件
        missing_files = list(set(files_shouldbe) - set(files_existing))
        if len(missing_files) > 0:
            miss_days = ", ".join(missing_files)
            raise FileNotFoundError(f"the following files are missing: {miss_days}")

        return files_shouldbe

    def checkformat(self, f: str, df: pd.DataFrame):
        """
        查每个文件里的格式
        """

        # 查表头
        if not {"K", "D", "T", self.facname} == set(df.columns):
            raise KeyError(f"incorrect header format in file {f}")

        # 查文件内是否有重复stock code
        df_dd = df.drop_duplicates(subset=["K"])
        if len(df_dd) < len(df):
            raise ValueError(f"duplicate stock codes in file {f}")

        # 查文件内日期是否与文件名一致，是否有不同日期出现在一个文件
        dt = (f.split("/")[-1]).split(".")[0]
        try:
            assert max(df.D) == min(df.D) == dt
        except AssertionError:
            raise ValueError(
                f"multiple dates in same file/filename doesn't match D column in file {f}"
            )

        # 查文件内的股票代码是否都是上市的股票
        stk_code = set(df_dd["K"])
        if not set(stk_code) <= set(self.cnall[dt]["K"]):
            raise ValueError(f"stock code beyond cnall exists in file {f}")


class FactorEvaluator(FactorBacktest):
    """
    因子评估器，用于加载因子到评估器，评估并输出
    """

    def __init__(
            self,
            sd,
            ed,
            backtest_name,
            univ_name,
            fdays,
            rettype,
            directory_list,
            fwdret,
            fwdret1,
            raw: dict,
            output_path: str,
            task_num: str,
            conn=None,
    ):  # 这里的data应该是之前FactorLoader.container
        super(FactorEvaluator, self).__init__(
            sd,
            ed,
            backtest_name,
            univ_name,
            fdays,
            rettype,
            directory_list,
        )
        self.alldays = fqdatac.get_daysrange(self.sd, self.ed)
        self.raw = raw
        self.fwd = fwdret
        self.fwd1 = fwdret1
        self.conn = conn
        self.output_path = output_path
        self.task_num = task_num

    def initialize(self):
        self.fac_processing = False
        self.whole_processing = True
        self.iterfac = iter(self.raw.keys())
        self.all_report = None

    # 切换到下一个因子去评估，将整个因子的数据加载到容器
    def nextfactor(self):
        try:
            self.curfac = next(self.iterfac)
            self.curraw = Factor.from_dict(self.raw[self.curfac])
            self.fac_processing = not self.fac_processing
            # self.iterdays = iter(sorted(self.curraw.facdays))
        except StopIteration:
            self.whole_processing = not self.whole_processing

    # 将下一天的因子和fwdret加载到container
    def nextfile(self):
        try:
            dt = next(self.iterdays)
            self.ff = FactorFile(
                kdtv=self.curraw.hist_d[dt], fac_name=self.curfac, dt=dt
            )
            self.curfwdret = self.fwd[dt]
        except StopIteration:
            self.fac_processing = not self.fac_processing

    # 将前一天的因子文件存起来，后面可能会有用
    def save_lastfile(self):
        self.lastfile = self.ff

    # 评估因子文件
    @timer
    def evaluate(self):
        try:
            # assert self.fac_processing and self.whole_processing
            assert self.whole_processing
        except AssertionError:
            raise IndexError(
                "the whole_process status flag and factor status flag are not both true"
            )

        facind = FactorIndicator(self.curraw.hist, self.fwd, self.fwd1, self.sd, self.ed)
        self.ic = facind.ic()
        # self.hitrate= facind.hitrate() #
        self.hitrate, self.graph_hit_rate_by_group = facind.hitrate_re()  ##
        self.graph_hitrate = facind.graph_hitrate_re()  ##
        self.ret, self.ret_by_group = facind.ret()
        self.wealth, self.wealth_by_group = facind.wealth()
        self.graph_return = facind.grpah_ret()
        # self.turnover = facind.turnover() #
        self.turnover = facind.turnover_re()
        self.coverage = facind.coverage()
        self.serial_correlation = facind.serial_corr()
        self.summary = facind.get_summary()
        self.table_performance_by_group = facind.performace_by_group
        self.graph_coverage = facind.graph_coverage()
        self.graph_ic = facind.graph_ic()
        self.graph_ic_by_industry = facind.graph_ic_industry()
        # self.graph_hit_rate_by_group = facind.graph_hit_rate_10bin() #
        self.graph_turnover = facind.graph_turnover_monthly()
        # self.graph_hitrate = facind.graph_hitrate_monthly() ##
        self.curreport = self.sf_report()

    # 合并得到包含单因子所有评价信息的dict
    def sf_report(self):
        single_report = {
            "Summary": self.summary["LS"],  # long-short 组合的performance
            "Coverage": self.coverage,
            "ICs": self.ic,
            "Turnover": self.turnover["LS"],  ####
            "graph_Turnover": self.graph_turnover["LS"],
            "LSreturns": self.graph_return["LS"],
            "Wealth": self.wealth["LS"],
            "HitRate": self.hitrate["LS"],  ####
            "graph_HitRate": self.graph_hitrate["LS"],
            "SCs": self.serial_correlation,
            "top_Summary": self.summary["top"],  # top 组合的performance
            "top_Turnover": self.turnover["top"],  ####
            "graph_top_Turnover": self.graph_turnover["top"],
            "top_Returns": self.graph_return["top"],
            "top_Wealth": self.wealth["top"],
            "top_HitRate": self.hitrate["top"],  ####
            "graph_top_HitRate": self.graph_hitrate["top"],
            "bottom_Summary": self.summary["bottom"],  # bottom 组合的performance
            "bottom_Turnover": self.turnover["bottom"],  ####
            "graph_bottom_Turnover": self.graph_turnover["bottom"],
            "bottom_Returns": self.graph_return["bottom"],
            "bottom_Wealth": self.wealth["bottom"],
            "bottom_HitRate": self.hitrate["bottom"],  ####
            "graph_bottom_HitRate": self.graph_hitrate["bottom"],
            # 图表的内容
            "graph_coverage": self.graph_coverage,
            "graph_ic": self.graph_ic,  # IC 月度图，key：D
            "graph_ic_by_industry": self.graph_ic_by_industry,  # IC分行业图, key:month
            "graph_wealth_by_group": self.wealth_by_group,  # 因子分10组净值曲线图，key：bin，D
            "graph_ret_by_group": self.ret_by_group,  #### 因子分10组收益率数据，key：bin，D
            "graph_hit_rate_by_group": self.graph_hit_rate_by_group,  # 分10组胜率图，key：bin
            "table_performance_by_group": self.table_performance_by_group,  # 分10组收益率，夏普比率，波动率，最大回撤，
        }
        return single_report

    def fac_report(self):
        # 将所有因子的report合并到总的report里面
        if self.all_report == None:
            self.all_report = {key: [] for key in self.curreport}
            for key in self.curreport:
                self.all_report[key].append(self.curreport[key])
        else:
            for key in self.curreport:
                self.all_report[key].append(self.curreport[key])

    # 将评估完的因子进行输出
    def output(self):
        # 对sel.all_report 遍历key，将结果dataframe concat起来，得到所有因子的report
        dict_output = {}
        dict_output["corr_matrix"] = self.factor_corrmatrix().reset_index()
        for key in self.all_report:
            try:
                if key in [
                    "Summary",
                    "top_Summary",
                    "bottom_Summary",
                    "table_performance_by_group",
                ]:
                    dict_output[key] = pd.concat(self.all_report[key], axis=0)
                    dict_output[key] = dict_output[key].reset_index(drop=True)
                elif key in ["graph_wealth_by_group", "graph_ret_by_group"]:
                    dict_output[key] = reduce(
                        lambda x, y: pd.merge(x, y, on=["D", "bin"], how="outer"),
                        self.all_report[key],
                    )
                    dict_output[key] = dict_output[key].reset_index(drop=True)
                elif key in ["graph_hit_rate_by_group"]:
                    dict_output[key] = reduce(
                        lambda x, y: pd.merge(x, y, on=["bin"], how="outer"),
                        self.all_report[key],
                    )
                    dict_output[key] = dict_output[key].reset_index(drop=True)
                elif key in ["graph_ic_by_industry"]:
                    dict_output[key] = reduce(
                        lambda x, y: pd.merge(
                            x, y, on=["ind_code", "ind_name", "bin"], how="outer"
                        ),
                        self.all_report[key],
                    )
                    dict_output[key] = dict_output[key].reset_index(drop=True)
                else:
                    dict_output[key] = reduce(
                        lambda x, y: pd.merge(x, y, on=["D"], how="outer"),
                        self.all_report[key],
                    )
                    dict_output[key] = dict_output[key].reset_index(drop=True)
            except:
                traceback.print_exc()
                print(f"多因子报告合并: {key}合并失败")
        # 写到Excel
        wb = openpyxl.Workbook(write_only=True)
        for sheet_name, df in dict_output.items():
            try:
                if isinstance(df, pd.DataFrame):
                    ws = wb.create_sheet(title=sheet_name)
                    # 写入header
                    ws.append(df.columns.tolist())
                    # 写入每行数据
                    for row in df.values.tolist():
                        ws.append(row)
            except:
                traceback.print_exc()
                print(f"excel writer: {sheet_name} 写入失败")

        today = datetime.now().strftime("%Y%m%d")
        file_name = f"BT_factor_report_{today}_{self.task_num}_{self.univ_name}_{self.sd}_{self.ed}.xlsx"

        # # 输出到指定本地路径
        # local_dir = self.output_path + "/" + self.task_num
        # os.makedirs(local_dir, exist_ok=True)
        # wb.save(local_dir + "/" + file_name)

        # # 生成程序运行成功信号
        # try:
        #     wb = openpyxl.load_workbook(local_dir + "/" + file_name)
        # except Exception as e:
        #     raise Exception(f"The generated Excel file could not be read. Reason: {e}")

        # success_file = os.path.join(local_dir, "success.txt")
        # with open(success_file, "w") as f:
        #     print(f"Generated {success_file}")    

        # 输出到指定路径
        try:
            output_dir = self.output_dir + "/" + self.task_num
            os.makedirs(output_dir, exist_ok=True)
            wb.save(output_dir + "/" + file_name)
            # 生成程序运行成功信号
            try:
                wb = openpyxl.load_workbook(output_dir + "/" + file_name)
            except Exception as e:
                raise Exception(f"The generated Excel file could not be read. Reason: {e}")
            success_file = os.path.join(output_dir, "success.txt")
            with open(success_file, "w") as f:
                print(f"Generated {success_file}")
        except:
            traceback.print_exc()


    # 因子相关性
    @timer
    def factor_corrmatrix(self):
        factor_names = list(self.raw.keys())
        dates = list(self.raw[factor_names[0]].keys())
        corr_matrix = pd.DataFrame(index=factor_names, columns=factor_names)
        for i in range(len(factor_names)):
            for j in range(i + 1, len(factor_names)):
                # 存储每日相关系数
                daily_corrs = []
                for d in dates:
                    # 如果对应日期的数据不存在，则返回nan
                    if d not in self.raw[factor_names[i]].keys() or d not in self.raw[factor_names[j]].keys():
                        daily_corrs.append(np.nan)
                        continue
                    fac1 = self.raw[factor_names[i]][d]
                    fac2 = self.raw[factor_names[j]][d]
                    # 保留具有相同股票的数据
                    merged = pd.merge(fac1, fac2, on=["K", "D", "T"], how="inner")
                    # 获取两个因子的列名
                    fac1_name = list(fac1.columns)[-1]
                    fac2_name = list(fac2.columns)[-1]
                    # 去除空值
                    merged.dropna(subset=[fac1_name, fac2_name], inplace=True)
                    # 计算每日相关系数
                    daily_corr = merged[fac1_name].corr(merged[fac2_name])
                    # 添加到列表
                    daily_corrs.append(daily_corr)
                # 求列表的平均值
                mean_corr = pd.Series(daily_corrs).mean().round(4)
                # 填入矩阵
                corr_matrix.loc[factor_names[i], factor_names[j]] = mean_corr
                corr_matrix.loc[factor_names[j], factor_names[i]] = mean_corr
            corr_matrix.loc[factor_names[i], factor_names[i]] = 1.0000

        return corr_matrix


class FactorIndicator:
    """
    factor indicators and calculation
    """

    def __init__(self, fac, fwd, fwd1, sd, ed):
        self.fac = fac
        self.fwd = fwd
        self.fwd1 = fwd1
        self.sd = sd
        self.ed = ed
        self.fac_name = list(self.fac.columns)[-1]
        self.fwd_ret_name = list(self.fwd.columns)[-1]
        self.fwd_ret_name1 = list(self.fwd1.columns)[-1]
        self.fac_fwd = self.fac_name + "_" + self.fwd_ret_name
        # 获取画图的参数：
        # 回测期一年以内为周，平均值rolling window 为4，
        # 超过一年为月，平均值rolling window 为6
        self.freq, self.window = self.get_graph_params()
        # self.fac_bin = FactorIndicator.bin_division(self.fac.copy(), 10)
        self.fac_bin = FactorIndicator.bin_division_re(self.fac.copy(), 10)

    # 获取画图groupby的参数
    def get_graph_params(self):
        d1 = pd.Period(self.sd + "M")
        d2 = pd.Period(self.ed + "M")
        n_month = pd.Series(pd.period_range(d1, d2, freq="M")).nunique()
        if n_month > 12:
            freq = "M"
            window = 6
        else:
            freq = "W-FRI"
            window = 4
        return freq, window

    # 判断本月有多少天
    def days_in_month(self, year, month):
        if month == 12:
            next_year = year + 1
            next_month = 1
        else:
            next_year = year
            next_month = month + 1
        end_of_month = datetime(next_year, next_month, 1) - datetime(year, month, 1)
        return (end_of_month).days

    # 判断日期间隔
    def get_date_interval(self, dt_start, dt_end):
        date_format = "%Y%m%d"

        # 将字符串日期转换为datetime对象
        start_date = datetime.strptime(dt_start, date_format)
        end_date = datetime.strptime(dt_end, date_format)

        # 计算日期间隔
        delta = end_date - start_date

        # 判断间隔
        if delta.days < self.days_in_month(start_date.year, start_date.month):
            return "day"
        elif delta.days < 365:
            return "week"
        else:
            return "month"

    @timer
    @staticmethod
    def bin_division(df: pd.DataFrame, nbins: int = 10):
        """
        将因子值分为x个bin
        结果的bin列范围为1~nbins，bin=1代表因子值最小组，bin=nbins代表因子值最大组
        """
        value_col = list(df.columns)[-1]
        df_dropna = df.copy().dropna(subset=[value_col])
        grouped = df_dropna.groupby(["D"])
        result_df = pd.DataFrame()
        for name, sub_df in grouped:
            sorted_df = sub_df.sort_values(by=value_col, ascending=True)
            sorted_df["rank"] = sorted_df.reset_index(drop=True).index + 1
            max_rank = max(sorted_df["rank"])
            sorted_df["rank_div_max"] = sorted_df["rank"] / max_rank
            try:
                assert nbins >= 1
                assert nbins <= max_rank
            except AssertionError:
                raise ValueError(
                    f"The range of nbins needs to be [1, length of the df]"
                )
            sorted_df["bin"] = ""
            for i in range(1, nbins + 1):
                bin_min = (i - 1) / float(nbins)
                bin_max = i / float(nbins)
                # 使用 .loc 来定位需要赋值的位置
                sorted_df.loc[
                    (sorted_df["rank_div_max"] > bin_min)
                    & (sorted_df["rank_div_max"] <= bin_max),
                    "bin",
                ] = i
            result_df = pd.concat([sorted_df, result_df])

        result_df.sort_values(by=["D", "bin"], ascending=True, inplace=True)
        return result_df[list(df.columns) + ["bin"]]

    @timer
    @staticmethod
    def bin_division_re(df: pd.DataFrame, nbins: int = 10):
        """
        将因子值分为x个bin
        结果的bin列范围为1~nbins，bin=1代表因子值最小组，bin=nbins代表因子值最大组
        """
        value_col = list(df.columns)[-1]
        result_df = df.dropna(subset=[value_col])
        result_df['rank'] = result_df.groupby('D')[value_col].rank(method='first')
        result_df['bin'] = result_df.groupby('D')['rank'].transform(lambda x: pd.qcut(x, nbins, labels=range(1,nbins+1,1)))
        result_df['bin'] = result_df['bin'].astype('int')
        result_df = result_df.drop(columns=["rank"])
        return result_df


    @timer
    @staticmethod
    def bin_division_duckdb(df: pd.DataFrame, nbins: int = 10):
        """
        将因子值分为x个bin
        结果的bin列范围为1~nbins，bin=1代表因子值最小组，bin=nbins代表因子值最大组
        """
        value_col = list(df.columns)[-1]
        df_dropna = df.dropna(subset=[value_col])
        con = duckdb.connect()
        con.register('df', df_dropna)
        sql = f"""
                WITH daily_ranked AS (
                SELECT 
                    *,
                    ROW_NUMBER() OVER (PARTITION BY "D" ORDER BY {value_col}) AS daily_rank 
                FROM df
                ),

                final AS (
                SELECT 
                    *,  COUNT(*) OVER (PARTITION BY "D") AS total_rows
                FROM daily_ranked 
                )

                SELECT
                *,
                CEIL(daily_rank * {nbins} / total_rows) AS bin
                FROM final
                """
        result_df = con.execute(sql).fetchdf()
        con.close()
        result_df['bin'] = result_df['bin'].astype(int)
        result_df.sort_values(by=["D", "bin"], ascending=True, inplace=True)
        return result_df[list(df.columns) + ["bin"]]


    # 定义一个resample函数，将时间序列按照给定频率的交易日末进行转换
    @staticmethod
    def df_resample(df, freq, method):
        # freq = "M","W-FRI"
        df.index = pd.to_datetime(df["D"], format="%Y%m%d")
        D_K = sorted(list(set(df.columns) & set(["D", "K"])))
        s_date = df.resample(freq).last()[D_K]
        if method == "mean":
            df = df.resample(freq).mean().reset_index(drop=True)
        elif method == "last":
            df = df.resample(freq).last().reset_index(drop=True)
        else:
            pass
        df[D_K] = s_date.values
        return df

    # 计算IC
    @timer
    def ic(self):
        fac = self.fac.dropna(subset=[self.fac_name])
        fwd = self.fwd.dropna(subset=[self.fwd_ret_name])
        merged = pd.merge(fac, fwd, how="inner", on=["K", "D"])

        def ic_group(group: pd.DataFrame) -> float:
            """
            计算T日的ic值
            """
            fac_rank = group[self.fac_name].rank(method="first")
            fwd_rank = group[self.fwd_ret_name].rank(method="first")
            rank_diff = fac_rank - fwd_rank
            n = len(group)
            if n <= 1:
                res = np.nan
            else:
                res = 1 - 6 * sum(rank_diff * rank_diff) / n / (n * n - 1)
            return res

        # 计算每天的IC值
        df_ic = merged.groupby("D").apply(ic_group).reset_index()
        df_ic.columns = ["D", self.fac_fwd]
        self.df_ic = df_ic
        return self.df_ic

    # 计算胜率
    @timer
    def hitrate(self):
        # 分组
        fac_bin = self.fac_bin.copy()
        # 合并因子值和收益率
        df_merge = pd.merge(fac_bin, self.fwd, on=["K", "D"], how="left")
        # 统计rk>0,rs<0,rk,rs,rk+rs的个数
        df_result = (
            df_merge.groupby("D")
            .agg(
                rs_0_cnt=(
                    self.fwd_ret_name,
                    lambda x: ((df_merge["bin"] == 1) & (x < 0)).sum(),
                ),
                rk_0_cnt=(
                    self.fwd_ret_name,
                    lambda x: ((df_merge["bin"] == 10) & (x > 0)).sum(),
                ),
                rs_cnt=("bin", lambda x: (x == 1).sum()),
                rk_cnt=("bin", lambda x: (x == 10).sum()),
                rk_rs_cnt=("bin", lambda x: ((x == 1) | (x == 10)).sum()),
            )
            .reset_index()
        )
        # 计算 hit rate
        df_result["LS"] = (df_result["rs_0_cnt"] + df_result["rk_0_cnt"]) / df_result[
            "rk_rs_cnt"
        ]
        df_result["top"] = df_result["rk_0_cnt"] / df_result["rk_cnt"]
        df_result["bottom"] = df_result["rs_0_cnt"] / df_result["rs_cnt"]
        # 构建结果字典
        self.dict_result_hitrate = {
            "LS": df_result[["D", "LS"]].rename(columns={"LS": self.fac_fwd}),
            "top": df_result[["D", "top"]].rename(columns={"top": self.fac_fwd}),
            "bottom": df_result[["D", "bottom"]].rename(
                columns={"bottom": self.fac_fwd}
            ),
        }

        return self.dict_result_hitrate

    @timer
    def hitrate_re(self):
        # 分组
        fac_bin = self.fac_bin.copy()
        # 合并因子值和收益率
        df_merge = pd.merge(fac_bin, self.fwd, on=["K", "D"], how="left")
        df_merge[self.fwd_ret_name] = df_merge[self.fwd_ret_name] > 0
        df_result = df_merge.groupby(by=["bin", "D"]).sum()[[self.fwd_ret_name]]
        df_result.columns = ["win"]
        df_result["count"] = df_merge.groupby(by=["bin", "D"]).count()[
            [self.fwd_ret_name]
        ]
        df_result["lose"] = df_result["count"] - df_result["win"]
        df_result = df_result.reset_index()
        # top 组合的hitrate
        top_hitrate = df_result[df_result["bin"] == 10].copy()
        top_hitrate[self.fac_fwd] = (
                top_hitrate.loc[:, "win"] / top_hitrate.loc[:, "count"]
        )
        # bottom 组合的hitrate
        bottom_hitrate = df_result[df_result["bin"] == 1].copy()
        bottom_hitrate.loc[:, self.fac_fwd] = (
                bottom_hitrate.loc[:, "lose"] / bottom_hitrate.loc[:, "count"]
        )
        # longshort 组合的hitrate
        top_df = top_hitrate[["D", "win", "lose", "count"]]
        top_df.columns = ["D", "win_top", "lose_top", "count_top"]
        bottom_df = bottom_hitrate[["D", "win", "lose", "count"]]
        bottom_df.columns = ["D", "win_bottom", "lose_bottom", "count_bottom"]
        ls_hitrate = pd.merge(top_df, bottom_df, how="left", on="D")
        ls_hitrate["ls_win"] = ls_hitrate["win_top"] + ls_hitrate["lose_bottom"]
        ls_hitrate["ls_count"] = ls_hitrate["count_top"] + ls_hitrate["count_bottom"]
        ls_hitrate.loc[:, self.fac_fwd] = (
                ls_hitrate.loc[:, "ls_win"] / ls_hitrate.loc[:, "ls_count"]
        )
        # 组合结果到dict
        columns = ["D", self.fac_fwd]
        ls_hitrate = ls_hitrate[columns].reset_index(drop=True)
        top_hitrate = top_hitrate[columns].reset_index(drop=True)
        bottom_hitrate = bottom_hitrate[columns].reset_index(drop=True)
        self.dict_result_hitrate = {
            "LS": ls_hitrate,
            "top": top_hitrate,
            "bottom": bottom_hitrate,
        }
        # 分10组的胜率
        df_result.loc[:, self.fac_fwd] = (
                df_result.loc[:, "win"] / df_result.loc[:, "count"]
        )
        hit_rate_by_group = (
            df_result.groupby(by="bin").mean()[[self.fac_fwd]].reset_index()
        )
        return self.dict_result_hitrate, hit_rate_by_group

    @timer
    def graph_hitrate_re(self):
        dict_hitrate = self.dict_result_hitrate.copy()
        # 计算到对应频率
        ls_hitrate = self.df_resample(dict_hitrate["LS"].copy(), self.freq, "mean")
        top_hitrate = self.df_resample(dict_hitrate["top"].copy(), self.freq, "mean")
        bottom_hitrate = self.df_resample(
            dict_hitrate["bottom"].copy(), self.freq, "mean"
        )
        # 计算平均线
        ls_hitrate[f"{self.fac_fwd}_avg"] = (
            ls_hitrate[self.fac_fwd].rolling(self.window).mean()
        )
        top_hitrate[f"{self.fac_fwd}_avg"] = (
            top_hitrate[self.fac_fwd].rolling(self.window).mean()
        )
        bottom_hitrate[f"{self.fac_fwd}_avg"] = (
            bottom_hitrate[self.fac_fwd].rolling(self.window).mean()
        )
        return {"LS": ls_hitrate, "top": top_hitrate, "bottom": bottom_hitrate}

    # 计算分组日度收益率
    @timer
    def ret(self):
        df_fac = self.fac_bin.copy()
        df_merge = pd.merge(df_fac, self.fwd1, on=["K", "D"], how="left")
        df_merge = df_merge.rename(columns={self.fwd_ret_name1: self.fac_fwd})
        # 计算分组收益率：日度
        self.ret_all = (
            df_merge[["D", "bin", self.fac_fwd]]
            .groupby(by=["D", "bin"])
            .mean()[self.fac_fwd]
            .reset_index()
        )
        self.ret_all = pd.pivot_table(
            self.ret_all, index="D", columns="bin", values=self.fac_fwd
        )
        self.ret_all = self.ret_all.shift(1).fillna(0).unstack().reset_index()
        self.ret_all.columns = ["bin", "D", self.fac_fwd]
        # 第1组的股票收益率：日度
        self.rs = self.ret_all[self.ret_all["bin"] == 1][["D", self.fac_fwd]]
        # 第10组的股票收益率：日度
        self.rk = self.ret_all[self.ret_all["bin"] == 10][["D", self.fac_fwd]]
        # Long Short 组合股票收益率：日度
        self.LSreturn = (
                self.rk.set_index("D", drop=True) - self.rs.set_index("D", drop=True)
        ).reset_index()
        return {"LS": self.LSreturn, "top": self.rk, "bottom": self.rs}, self.ret_all

    # 计算分组净值图，LS, top， bottom的净值图
    @timer
    def wealth(self):
        # 计算所有组合的净值
        self.wealth_by_group = self.ret_all.sort_values(by="D", ascending=True).copy()
        self.wealth_by_group = pd.pivot_table(
            self.wealth_by_group, columns="bin", index="D", values=self.fac_fwd
        )
        self.wealth_by_group = (
            (self.wealth_by_group + 1).cumprod().unstack().reset_index()
        )
        self.wealth_by_group.columns = ["bin", "D", self.fac_fwd]
        # 计算LS组合的净值
        self.df_wealth = self.LSreturn.copy()
        self.df_wealth[self.fac_fwd] = (self.df_wealth[self.fac_fwd] + 1).cumprod()
        # 计算top组合的净值
        self.top_wealth = self.rk.copy()
        self.top_wealth[self.fac_fwd] = (self.top_wealth[self.fac_fwd] + 1).cumprod()
        # 计算bottom组合的净值
        self.bottom_wealth = self.rs.copy()
        self.bottom_wealth[self.fac_fwd] = (
                self.bottom_wealth[self.fac_fwd] + 1
        ).cumprod()
        return {
            "LS": self.df_wealth,
            "top": self.top_wealth,
            "bottom": self.bottom_wealth,
        }, self.wealth_by_group

    # 计算单个组合的因子收益率图：周/月
    def sub_graph_ret(self, wealth, freq, method):
        # 转化净值为目标输出频率
        df_ret_freq = self.df_resample(wealth.copy(), freq, method)
        df_ret_freq = pd.concat([wealth.iloc[[0], :], df_ret_freq], axis=0).reset_index(
            drop=True
        )
        # 计算目标输出频率的因子收益率
        df_ret_freq[self.fac_fwd] = df_ret_freq[self.fac_fwd].pct_change()
        df_ret_freq = df_ret_freq.dropna()
        df_ret_freq[f"{self.fac_fwd}_avg"] = (
            df_ret_freq[self.fac_fwd].rolling(self.window).mean()
        )
        return df_ret_freq

    # 计算LS, top， bottom的因子收益率图:周/月
    @timer
    def grpah_ret(self):
        df_ret_freq = self.sub_graph_ret(self.df_wealth.copy(), self.freq, "last")
        top_ret_freq = self.sub_graph_ret(self.top_wealth.copy(), self.freq, "last")
        bottom_ret_freq = self.sub_graph_ret(
            self.bottom_wealth.copy(), self.freq, "last"
        )
        return {"LS": df_ret_freq, "top": top_ret_freq, "bottom": bottom_ret_freq}

    # 计算换手、序列相关性
    @timer
    def turnover(self):
        # 分成10组
        fac_bin = self.fac_bin.copy()
        # 获取日期列表
        dt_list = list(sorted(set(fac_bin["D"])))

        result_LS, result_top, result_bottom = [], [], []
        prev_s1_set, prev_s10_set = set(), set()

        for dt in dt_list:
            # 如果为第一天，换手率为1
            if dt == dt_list[0]:
                result_LS.append([dt, 1])
                result_top.append([dt, 1])
                result_bottom.append([dt, 1])
                # 第一日的股票集合，给第二天用
                prev_s1_set = set(
                    fac_bin[(fac_bin["D"] == dt) & (fac_bin["bin"] == 1)]["K"]
                )
                prev_s10_set = set(
                    fac_bin[(fac_bin["D"] == dt) & (fac_bin["bin"] == 10)]["K"]
                )
                continue

            dt_1 = fqdatac.get_day(dt, -1)

            # 当前日期的股票集合
            s1_set = set(fac_bin[(fac_bin["D"] == dt) & (fac_bin["bin"] == 1)]["K"])
            s10_set = set(fac_bin[(fac_bin["D"] == dt) & (fac_bin["bin"] == 10)]["K"])

            # T-1日的股票集合
            f1_set = prev_s1_set
            c1 = len(s1_set & f1_set)
            f10_set = prev_s10_set
            c10 = len(s10_set & f10_set)

            prev_s1_set, prev_s10_set = s1_set, s10_set

            if s1_set or s10_set:
                result_LS.append([dt, 1 - (c1 + c10) / (len(s1_set) + len(s10_set))])
                result_top.append([dt, 1 - c10 / len(s10_set)])
                result_bottom.append([dt, 1 - c1 / len(s1_set)])
            else:
                result_LS.append([dt, 0])
                result_top.append([dt, 0])
                result_bottom.append([dt, 0])

        df_LS = pd.DataFrame(result_LS, columns=["D", self.fac_fwd])
        df_top = pd.DataFrame(result_top, columns=["D", self.fac_fwd])
        df_bottom = pd.DataFrame(result_bottom, columns=["D", self.fac_fwd])
        self.dict_turnover = {"LS": df_LS, "top": df_top, "bottom": df_bottom}
        return self.dict_turnover

    @timer
    def turnover_re(self):
        # 分成10组
        fac_bin = self.fac_bin.copy()
        df_turnover = (
            fac_bin.groupby(by=["bin", "D"])["K"]
            .apply(lambda x: set(x.tolist()))
            .to_frame()
        )
        df_turnover.columns = ["K_set"]
        df_turnover = df_turnover.reset_index()
        df_turnover["K_count"] = df_turnover["K_set"].apply(lambda x: len(x))
        df_turnover["prev_K_set"] = (
            df_turnover.groupby("bin")["K_set"].apply(lambda x: x.shift(1)).values
        )
        df_turnover["overlap"] = df_turnover.apply(
            lambda x: x["K_set"] & x["prev_K_set"]
            if isinstance(x["K_set"], set) and isinstance(x["prev_K_set"], set)
            else 0,
            axis=1,
        )
        df_turnover["n_overlap"] = df_turnover["overlap"].apply(
            lambda x: len(x) if isinstance(x, set) else x
        )
        df_turnover[self.fac_fwd] = (
                1 - df_turnover["n_overlap"] / df_turnover["K_count"]
        )
        # 分10组的turnover
        self.turnover_by_group = df_turnover[["bin", "D", self.fac_fwd]]
        # top和bottom组合的turnover
        top_turnover = df_turnover[df_turnover["bin"] == 10][["D", self.fac_fwd]]
        bottom_turnover = df_turnover[df_turnover["bin"] == 1][["D", self.fac_fwd]]
        # longshort 组合的turnover
        cols = ["K_set", "K_count", "prev_K_set", "overlap", "n_overlap"]
        top_df = df_turnover[df_turnover["bin"] == 10][["D"] + cols]
        top_df.columns = ["D"] + [i + "_top" for i in cols]
        bottom_df = df_turnover[df_turnover["bin"] == 1][["D"] + cols]
        bottom_df.columns = ["D"] + [i + "_bottom" for i in cols]
        ls_turnover = pd.merge(top_df, bottom_df, how="left", on="D")
        ls_turnover["t_overlap"] = (
                ls_turnover["n_overlap_top"] + ls_turnover["n_overlap_bottom"]
        )
        ls_turnover["t_count"] = (
                ls_turnover["K_count_top"] + ls_turnover["K_count_bottom"]
        )
        ls_turnover[self.fac_fwd] = (
                1 - ls_turnover["t_overlap"] / ls_turnover["t_count"]
        )
        # self.dict_df_turnover = {"LS":ls_turnover.copy(),
        #                          'top':df_turnover[df_turnover["bin"] == 10],
        #                          'bottom':df_turnover[df_turnover["bin"] == 1]}
        ls_turnover = ls_turnover[["D", self.fac_fwd]]
        self.dict_turnover = {
            "LS": ls_turnover,
            "top": top_turnover,
            "bottom": bottom_turnover,
        }
        return self.dict_turnover

    # 计算覆盖率
    @timer
    def coverage(self):
        df = self.fac.dropna(subset=[self.fac_name])
        df_coverage = df.groupby("D")["K"].count().reset_index()
        df_coverage.columns = ["D", self.fac_fwd]
        return df_coverage

    # 计算因子序列相关性
    @timer
    def serial_corr(self):
        temp_df = self.fac.sort_values(by="D", ascending=True).copy()
        # 获取前一天的因子序列
        prev_fac = pd.pivot_table(
            temp_df, columns="K", index="D", values=self.fac_name
        ).shift(1)
        prev_fac = prev_fac.unstack().reset_index()
        prev_fac.columns = ["K", "D", "prev_fac"]
        prev_fac = prev_fac.dropna(subset=["prev_fac"])
        # 滚动计算相关性
        temp_df = pd.merge(temp_df, prev_fac, how="left", on=["K", "D"])
        self.serial_correlation = temp_df.groupby("D").apply(
            lambda x: x[[self.fac_name, "prev_fac"]].corr(method="spearman").iloc[0, 1]
        )
        self.serial_correlation = (
            self.serial_correlation.to_frame().reset_index().dropna()
        )
        self.serial_correlation.columns = ["D", self.fac_fwd]
        return self.serial_correlation

    # 计算summary中：coverage_max coverage_min
    @timer
    def summary_coverage(self):
        fac_dropna = self.fac.dropna(subset=[self.fac_name])
        coverage_max = len(fac_dropna["K"].unique())
        total_days = len(self.fac["D"].unique())
        df_days = fac_dropna.groupby("K")["D"].count()
        coverage_min = len(df_days[df_days == total_days])
        df_coverage = pd.DataFrame(
            {"coverage_max": [coverage_max], "coverage_min": [coverage_min]},
            index=[self.fac_fwd],
        )
        return df_coverage

    # 计算summary中：rank_ic_mean rank_ic_std ic_ir
    @timer
    def summary_ic_ir(self):
        df_ic = self.df_ic.copy()
        ic_mean = df_ic[self.fac_fwd].describe()["mean"] * 100
        rank_ic_mean = round(ic_mean, 2)
        ic_std = df_ic[self.fac_fwd].describe()["std"] * 100
        rank_ic_std = round(ic_std, 2)
        ic_ir = ic_mean / ic_std
        ic_ir = round(ic_ir, 2)
        df_ic_ir = pd.DataFrame(
            {
                "rank_ic_mean": [rank_ic_mean],
                "rank_ic_std": [rank_ic_std],
                "ic_ir": [ic_ir],
            },
            index=[self.fac_fwd],
        )
        return df_ic_ir

    # 计算summary中：rank_ic_mean_year_YYYY
    @timer
    def summary_ic_year(self):
        df_ic = self.df_ic.copy()
        df_ic["year"] = df_ic["D"].str[:4].astype(int)
        df_ic = df_ic.groupby("year")[self.fac_fwd].mean() * 100
        df_ic = round(df_ic, 2).to_frame().T
        df_ic_year = df_ic.add_prefix("rank_ic_mean_year_")
        return df_ic_year

    # 返回LS、top、bottom三组summary_hit_rate的dict，包含3个df
    @timer
    def summary_hit_rate(self):
        dict_hitrate = self.dict_result_hitrate
        hitrate_LS = pd.DataFrame(
            [round(dict_hitrate["LS"][self.fac_fwd].mean() * 100, 2)],
            columns=["hit_rate"],
            index=[self.fac_fwd],
        )
        hitrate_top = pd.DataFrame(
            [round(dict_hitrate["top"][self.fac_fwd].mean() * 100, 2)],
            columns=["hit_rate"],
            index=[self.fac_fwd],
        )
        hitrate_bottom = pd.DataFrame(
            [round(dict_hitrate["bottom"][self.fac_fwd].mean() * 100, 2)],
            columns=["hit_rate"],
            index=[self.fac_fwd],
        )
        return {"LS": hitrate_LS, "top": hitrate_top, "bottom": hitrate_bottom}

    # 返回LS、top、bottom三组summary_turnover的dict，包含3个df
    @timer
    def summary_turnover(self):
        dict_turnover = self.dict_turnover
        turnover_LS = pd.DataFrame(
            [round(dict_turnover["LS"][self.fac_fwd].mean() * 100,2)],
            columns=["turnover"],
            index=[self.fac_fwd],
        )
        turnover_top = pd.DataFrame(
            [round(dict_turnover["top"][self.fac_fwd].mean() * 100,2)],
            columns=["turnover"],
            index=[self.fac_fwd],
        )
        turnover_bottom = pd.DataFrame(
            [round(dict_turnover["bottom"][self.fac_fwd].mean() * 100,2)],
            columns=["turnover"],
            index=[self.fac_fwd],
        )
        return {"LS": turnover_LS, "top": turnover_top, "bottom": turnover_bottom}

    # 返回因子coverage图数据：月均coverage 或 周均coverage
    @timer
    def graph_coverage(self):
        df_cv = self.coverage().copy()
        df_cv = self.df_resample(df_cv, self.freq, "mean")
        def custom_round(x):
            if pd.notnull(x):
                return int(round(x))
            else:
                return x
        df_cv[self.fac_fwd] = df_cv[self.fac_fwd].apply(lambda x: custom_round(x))
        return df_cv

    # 返回因子IC表现图数据：月均ic及其rolling_avg 或 周均ic及其rolling_avg
    @timer
    def graph_ic(self):
        df_ic = self.df_ic.copy()
        df_ic = self.df_resample(df_ic, self.freq, "mean")
        df_ic[self.fac_fwd + "_avg"] = (
            df_ic[self.fac_fwd].rolling(window=self.window).mean()
        )
        return df_ic

    # 返回IC行业分布图数据：ic_mean
    @timer
    def graph_ic_industry(self):
        fac = self.fac.dropna(subset=[self.fac_name])
        fwd = self.fwd.dropna(subset=[self.fwd_ret_name])
        merged = pd.merge(fac, fwd, how="inner", on=["K", "D"])
        sd = merged["D"].min()
        ed = merged["D"].max()
        bin_list = list(range(1, 31))
        df_ic_ind = pd.DataFrame(
            {
                "bin": bin_list,
                "ind_code": IND_DICT.keys(),
                "ind_name": IND_DICT.values(),
                self.fac_fwd: "",
            },
            index=IND_DICT.keys(),
        )
        # 获取每个行业的成分股并计算ic
        for k, v in IND_DICT.items():
            ind = fqdatac.get_indconst(sd, ed, ind_code=k)
            fac_ind = pd.merge(ind, merged, how="inner", on=["K", "D"])

            def ic_group(group: pd.DataFrame) -> float:
                """
                计算T日的ic值
                """
                fac_rank = group[self.fac_name].rank(method="first")
                fwd_rank = group[self.fwd_ret_name].rank(method="first")
                rank_diff = fac_rank - fwd_rank
                n = len(group)
                if n <= 1:
                    res = np.nan
                else:
                    res = 1 - 6 * sum(rank_diff * rank_diff) / n / (n * n - 1)
                return res

            # 计算该行业的ic_mean
            ic_ind = fac_ind.groupby("D").apply(ic_group).mean()
            df_ic_ind.loc[(df_ic_ind["ind_code"] == k), self.fac_fwd] = ic_ind
        return df_ic_ind

    # 因子分十组的胜率
    @timer
    def graph_hit_rate_10bin(self):
        # 分组
        fac_bin = self.fac_bin.copy()
        # 合并因子值和收益率
        df_merge = pd.merge(fac_bin, self.fwd, on=["K", "D"], how="left")
        # 统计10组的r
        df_result = (
            df_merge.groupby("D")
            .agg(
                **{
                    f"bin{i}_0_cnt": (
                        self.fwd_ret_name,
                        lambda x, i=i: ((df_merge["bin"] == i) & (x > 0)).sum(),
                    )
                    for i in range(1, 11)
                },
                **{
                    f"bin{i}_cnt": ("bin", lambda x, i=i: (x == i).sum())
                    for i in range(1, 11)
                },
            )
            .reset_index()
        )
        # 计算每个 bin 的 hitrate
        bin_columns = [f"bin{i}" for i in range(1, 11)]
        df_result[bin_columns] = df_result[[f"bin{i}_0_cnt" for i in range(1, 11)]].div(
            df_result[[f"bin{i}_cnt" for i in range(1, 11)]].values
        )
        # 计算每个 bin 的均值
        df_hitrate = df_result[bin_columns].mean()
        df_hitrate = pd.DataFrame(
            {"bin": [i for i in range(1, 11)], self.fac_fwd: df_hitrate}
        )

        return df_hitrate

    # 单因子的月度换手率
    # 短于一个月返回日度，短于一年返回周度，长于一年返回月度
    @timer
    def graph_turnover_monthly(self):
        interval = self.get_date_interval(self.sd, self.ed)
        return self.turnover_different_freq(interval)

    @timer
    def graph_hitrate_monthly(self):
        interval = self.get_date_interval(self.sd, self.ed)
        return self.hitrate_different_freq(interval)

    def hitrate_different_freq(self, interval: str):
        hit_dict = self.hitrate()
        if interval == "day":
            roll = 5
        elif interval == "week":
            roll = 4
        else:
            roll = 6
        for k in list(hit_dict.keys()):
            df = hit_dict[k].copy()
            if interval != "day":
                df["D"] = pd.to_datetime(df["D"])  # 将'D'列转换为日期时间格式
                df.set_index("D", inplace=True)  # 设置'D'列为索引
                if interval == "week":
                    df = df.resample("W-FRI").mean()
                elif interval == "month":
                    df = df.resample("M").mean()
                df.reset_index(inplace=True)
                df.dropna(inplace=True)
                df["D"] = df["D"].dt.strftime("%Y%m%d")
            df[f"{self.fac_fwd}_avg"] = df[self.fac_fwd].rolling(roll).mean()
            hit_dict[k] = df
        return hit_dict

    def turnover_different_freq(self, interval: str):
        fac_bin = self.fac_bin.copy()
        bin_set = fac_bin.groupby(["D", "bin"])["K"].apply(set).reset_index()
        if interval != "day":
            bin_set["D"] = pd.to_datetime(bin_set["D"])  # 将'D'列转换为日期时间格式
            result = {}
            for bin_value, group in bin_set.groupby("bin"):
                group.set_index("D", inplace=True)  # 设置'D'列为索引
                if interval == "week":
                    resampled_last = group.resample("W-FRI").last()
                elif interval == "month":
                    resampled_last = group.resample("M").last()
                result[bin_value] = resampled_last
            fac_bin_resp = pd.DataFrame()
            for l in list(result.values()):
                fac_bin_resp = pd.concat([fac_bin_resp, l.reset_index()])
            fac_bin_resp["D"] = fac_bin_resp["D"].dt.strftime("%Y%m%d")
        else:
            fac_bin_resp = bin_set
        fac_bin_resp.dropna(inplace=True)
        # 获取日期列表
        dt_list = list(sorted(set(fac_bin_resp["D"])))
        result_LS, result_top, result_bottom = [], [], []
        prev_s1_set, prev_s10_set = set(), set()
        for dt in dt_list:
            # 如果为第一天，换手率为1
            if dt == dt_list[0]:
                result_LS.append([dt, 1])
                result_top.append([dt, 1])
                result_bottom.append([dt, 1])
                # 第一日的股票集合，给第二天用
                prev_s1_set = fac_bin_resp[
                    (fac_bin_resp["D"] == dt) & (fac_bin_resp["bin"] == 1)
                    ]["K"].iloc[0]
                prev_s10_set = fac_bin_resp[
                    (fac_bin_resp["D"] == dt) & (fac_bin_resp["bin"] == 10)
                    ]["K"].iloc[0]
                continue
            # dt_1 = dt_list[dt_list.index(dt)-1]
            # 当前日期的股票集合
            s1_set = fac_bin_resp[
                (fac_bin_resp["D"] == dt) & (fac_bin_resp["bin"] == 1)
                ]["K"].iloc[0]
            s10_set = fac_bin_resp[
                (fac_bin_resp["D"] == dt) & (fac_bin_resp["bin"] == 10)
                ]["K"].iloc[0]
            # T-1日的股票集合
            f1_set = prev_s1_set
            c1 = len(s1_set & f1_set)
            f10_set = prev_s10_set
            c10 = len(s10_set & f10_set)
            prev_s1_set, prev_s10_set = s1_set, s10_set
            if s1_set or s10_set:
                result_LS.append([dt, 1 - (c1 + c10) / (len(s1_set) + len(s10_set))])
                result_top.append([dt, 1 - c10 / len(s10_set)])
                result_bottom.append([dt, 1 - c1 / len(s1_set)])
            else:
                result_LS.append([dt, 0])
                result_top.append([dt, 0])
                result_bottom.append([dt, 0])
        df_LS = pd.DataFrame(result_LS, columns=["D", self.fac_fwd])
        df_top = pd.DataFrame(result_top, columns=["D", self.fac_fwd])
        df_bottom = pd.DataFrame(result_bottom, columns=["D", self.fac_fwd])
        if interval == "day":
            roll = 5
        elif interval == "week":
            roll = 4
        else:
            roll = 6
        df_LS[f"{self.fac_fwd}_avg"] = df_LS[self.fac_fwd].rolling(roll).mean()
        df_top[f"{self.fac_fwd}_avg"] = df_top[self.fac_fwd].rolling(roll).mean()
        df_bottom[f"{self.fac_fwd}_avg"] = df_bottom[self.fac_fwd].rolling(roll).mean()
        return {"LS": df_LS, "top": df_top, "bottom": df_bottom}

    # 计算持有期收益率
    def get_hpr(self, df_wealth, days, annualize):
        if days > len(df_wealth):
            temp_wealth = df_wealth[self.fac_fwd].copy()
        else:
            temp_wealth = df_wealth[self.fac_fwd].iloc[-days:]
        hpr = temp_wealth.iloc[-1] / temp_wealth.iloc[0] - 1
        if annualize:
            # n_year = len(temp_wealth) / 252
            # hpr = (1 + hpr) ** (1 / n_year) - 1
            daily_return = temp_wealth.pct_change()
            hpr = daily_return.mean() * (252**0.5)
            return hpr
        else:
            return hpr

    # 计算最大回撤
    def maxdrawdown(self, df_wealth):
        temp_wealth = df_wealth.sort_values(by="D", ascending=True).copy()
        temp_wealth.loc[:, "cummax"] = (
            temp_wealth[self.fac_fwd].to_frame().expanding().max()[self.fac_fwd]
        )
        temp_wealth.loc[:, "drawdown"] = (
                                                 temp_wealth["cummax"] - temp_wealth[self.fac_fwd]
                                         ) / temp_wealth["cummax"]
        max_drawdown = temp_wealth.loc[:, "drawdown"].max()
        return max_drawdown

    # 计算综合performance
    def ret_performance(self, df_wealth, df_return):
        # return_1m: 回测期最后21个交易日的return（%）
        return_1m = self.get_hpr(df_wealth=df_wealth, days=21, annualize=False)
        # return_3m: 回测期最后63个交易日的return（%）
        return_3m = self.get_hpr(df_wealth=df_wealth, days=53, annualize=False)
        # cagr_1y: 回测期最后252个交易日的return（%）
        cagr_1y = self.get_hpr(df_wealth=df_wealth, days=252, annualize=False)
        # cagr_3y: 回测期最后回测期最后756个交易日的年化return（%）
        cagr_3y = self.get_hpr(df_wealth=df_wealth, days=756, annualize=True)
        # cagr_5y: 回测期最后回测期最后1260个交易日的年化return（%）
        cagr_5y = self.get_hpr(df_wealth=df_wealth, days=1260, annualize=True)
        # cagr: 回测期的return（%)
        cagr = self.get_hpr(df_wealth=df_wealth, days=len(df_wealth), annualize=True)
        # volatility: 回测期因子组合收益的波动率（%）
        volatility = df_return[self.fac_fwd].std() * np.sqrt(252)
        # sharpe_ratio:回测期因子组合的夏普比率, 暂时用rf = 0.02
        sharp_ratio = (cagr - 0.02) / volatility
        # maximum drawdown:回测期因子组合的最大回撤
        max_drawdown = self.maxdrawdown(df_wealth)
        performance = pd.DataFrame(
            {
                "return_1m": round(return_1m * 100, 2),
                "return_3m": round(return_3m * 100, 2),
                "cagr_1y": round(cagr_1y * 100, 2),
                "cagr_3y": round(cagr_3y * 100, 2),
                "cagr_5y": round(cagr_5y * 100, 2),
                "cagr": round(cagr * 100, 2),
                "volatility": round(volatility * 100, 2),
                "sharp_ratio": round(sharp_ratio, 2),
                "max_drawdown": round(max_drawdown * 100, 2),
            },
            index=[self.fac_fwd],
        )
        return performance

    # 返回LS、top、bottom三组summary_ret的1月，3月，1年，3年，5年，收益率，波动率，sharp ratio
    @timer
    def summary_ret(self):
        # LS 组合的performance
        ls_performance = self.ret_performance(self.df_wealth, self.LSreturn)
        # top 组合的performance
        top_performance = self.ret_performance(self.top_wealth, self.rk)
        # bottom 组合的performance
        bottom_performance = self.ret_performance(self.bottom_wealth, self.rs)
        # 分10组的perfromance：
        group_perfromance = []
        for bin in set(self.wealth_by_group["bin"].to_list()):
            sub_wealth = self.wealth_by_group[self.wealth_by_group["bin"] == bin][
                ["D", self.fac_fwd]
            ]
            sub_ret = self.ret_all[self.ret_all["bin"] == bin][["D", self.fac_fwd]]
            sub_performance = self.ret_performance(sub_wealth, sub_ret)
            sub_performance.insert(loc=0, column="bin", value=[bin])
            sub_performance = sub_performance.reset_index(drop=True)
            sub_performance.insert(loc=0, column="fac_name", value=[self.fac_fwd])
            group_perfromance.append(sub_performance)
        group_perfromance = pd.concat(group_perfromance, axis=0)
        return {
            "LS": ls_performance,
            "top": top_performance,
            "bottom": bottom_performance,
        }, group_perfromance

    # 计算分年度cagr
    @timer
    def cagr_by_year(self, df_wealth):
        temp_wealth = df_wealth.copy().sort_values(by="D", ascending=True)
        temp_wealth["year"] = temp_wealth["D"].str.slice(0, 4)
        col_name = []
        ret_year = []
        for year in temp_wealth["year"].unique():
            sub_wealth = temp_wealth[temp_wealth["year"] == year].copy()
            ret_i = (
                    sub_wealth[self.fac_fwd].iloc[-1] / sub_wealth[self.fac_fwd].iloc[0] - 1
            )
            col_name.append(year)
            ret_year.append(round(ret_i * 100, 2))
        col_name = ["cagr_" + i for i in col_name]
        data = dict(zip(col_name, ret_year))
        df_cagr_by_year = pd.DataFrame(data, index=[self.fac_fwd])
        return df_cagr_by_year

    # 返回LS、top、bottom三组cagt_by_year
    @timer
    def summary_cagr_year_t(self):
        ls_cagr_by_year = self.cagr_by_year(self.df_wealth)
        top_cagr_by_year = self.cagr_by_year(self.top_wealth)
        bottom_cagr_by_year = self.cagr_by_year(self.bottom_wealth)
        return {
            "LS": ls_cagr_by_year,
            "top": top_cagr_by_year,
            "bottom": bottom_cagr_by_year,
        }

    # 计算平均serial correlation
    @timer
    def summary_sc(self):
        avg_sc = self.serial_correlation[self.fac_fwd].mean()
        summary_sc = pd.DataFrame(
            {"serial_corr": round(avg_sc * 100, 2)}, index=[self.fac_fwd]
        )
        return summary_sc

    # 获得LS，top，bottom的summary report
    @timer
    def get_summary(self):
        # coverage_max, coverage_min
        df_coverage = self.summary_coverage()

        # return_1m,return_3m,cagr_1y,cagr_3y,cagr_5y,cagr,volatility,sharpe_ratio,max_drawdown
        dict_ret, self.performace_by_group = self.summary_ret()

        # hit_rate
        dict_hit_rate = self.summary_hit_rate()

        # rank_ic_mean, rank_ic_std, ic_ir
        df_ic_ir = self.summary_ic_ir()

        # serial correlation
        df_sc = self.summary_sc()

        # turnover
        dict_turnover = self.summary_turnover()

        # cagr_year_t
        dict_cagr_year = self.summary_cagr_year_t()

        # rank_ic_mean_year: 列数量与回测区间包含的年度数量有关
        df_ic_year = self.summary_ic_year()

        summary = {}
        for key in ["LS", "top", "bottom"]:
            all_df = [
                df_coverage,
                dict_ret[key],
                dict_hit_rate[key],
                df_ic_ir,
                df_sc,
                dict_turnover[key],
                dict_cagr_year[key],
                df_ic_year,
            ]
            sub_summary = pd.concat(all_df, axis=1)
            sub_summary.insert(loc=0, column="backtest_ed", value=[self.ed])
            sub_summary.insert(loc=0, column="backtest_sd", value=[self.sd])
            sub_summary = sub_summary.reset_index(drop=True)
            sub_summary.insert(loc=0, column="fac_name", value=[self.fac_fwd])
            summary[key] = sub_summary
        return summary
