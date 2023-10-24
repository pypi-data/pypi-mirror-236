# -*- coding:utf-8 -*-
"""
常用的excel文件处理方法
"""
# 安装win32com：pip install pypiwin32
from copy import copy
from pathlib import Path
from typing import Union

from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

from .paths import make_valid_name, get_usable_path


def xls_to_xlsx(path: Union[str, Path], del_src: bool = True) -> Union[Path, bool]:
    """把xls格式另存为xlsx格式                  \n
    :param path: xls文件路径
    :param del_src: 另存之后是否删除源文件
    :return: 是否成功
    """
    from win32com import client
    excel = None
    try:
        excel = client.Dispatch("Ket.Application")  # 调用WPS程序
    except:
        try:
            excel = client.gencache.EnsureDispatch('Excel.Application')  # 调用Office程序
        except:
            pass

    if excel:
        path = str(Path(path).absolute())
        wb = excel.Workbooks.Open(path)
        new_name = get_usable_path(f"{path}x")
        wb.SaveAs(str(new_name), 51)
        wb.Close()
        if del_src:
            Path(path).unlink()
        return new_name

    return False


def copy_cell(cell_or_loc: Union[Cell, str],
              tar_sheet: Worksheet,
              row_or_loc: Union[str, int],
              col: Union[str, int] = None,
              src_sheet: Worksheet = None,
              copy_format: bool = True) -> None:
    """复制单元格到目标工作表                               \n
    :param cell_or_loc: 源单元格或其坐标（'B3'或'3,3'），输入坐标时必须输入src_sheet
    :param tar_sheet: 目标工作表
    :param row_or_loc: 目标工作表中单元格行号或位置字符串
    :param col: 目标工作表中单元格位置
    :param src_sheet: 源工作表
    :param copy_format: 是否复制格式
    :return: None
    """
    if isinstance(cell_or_loc, str):
        if not src_sheet:
            raise ValueError('用坐标定位时必须输入sheet')

        if ',' in cell_or_loc:
            loc = cell_or_loc.split(',')
            cell_or_loc = src_sheet.cell(int(loc[0].strip()), int(loc[1].strip()))
        else:
            cell_or_loc = src_sheet[cell_or_loc]

    if isinstance(row_or_loc, str):
        tar_cell = tar_sheet[row_or_loc]
    elif isinstance(row_or_loc, int) and isinstance(col, int):
        tar_cell = tar_sheet.cell(row_or_loc, col)
    elif isinstance(row_or_loc, int) and isinstance(col, str):
        tar_cell = tar_sheet[f'{col}{row_or_loc}']
    else:
        raise ValueError('传入的位置参数不正确')

    tar_cell.value = cell_or_loc.value
    tar_cell.hyperlink = cell_or_loc.hyperlink

    if copy_format and cell_or_loc.has_style:
        tar_cell.font = copy(cell_or_loc.font)
        tar_cell.border = copy(cell_or_loc.border)
        tar_cell.fill = copy(cell_or_loc.fill)
        tar_cell.number_format = copy(cell_or_loc.number_format)
        tar_cell.protection = copy(cell_or_loc.protection)
        tar_cell.alignment = copy(cell_or_loc.alignment)


def copy_row(row_or_num: Union[int, tuple],
             tar_sheet: Worksheet,
             src_sheet: Worksheet = None,
             tar_row_num: int = None,
             copy_col_width: bool = True) -> None:
    """复制一行到目标工作表                             \n
    :param row_or_num: 源行对象或行号
    :param tar_sheet: 目标工作表
    :param src_sheet: 源工作表
    :param tar_row_num: 目标行号，默认为在最后添加一行
    :param copy_col_width: 是否复制列宽
    :return: None
    """
    if not tar_row_num:
        max_row = tar_sheet.max_row
        if len(tar_sheet[max_row]) == 1 and tar_sheet[max_row][0].value is None:
            tar_row_num = 1
        else:
            tar_row_num = max_row + 1

    tar_row_num = tar_row_num or tar_sheet.max_row + 1
    if isinstance(row_or_num, tuple):
        src_sheet = row_or_num[0].parent
        row = row_or_num
        num = row[0].row
    else:
        if src_sheet is None:
            raise ValueError('输入行号必须同时输入源表格。')
        row = src_sheet[row_or_num]
        num = row_or_num

    tar_sheet.row_dimensions[tar_row_num].height = src_sheet.row_dimensions[num].height

    for key, cell in enumerate(row, 1):
        col = get_column_letter(key)
        if copy_col_width:
            tar_sheet.column_dimensions[col].width = src_sheet.column_dimensions[col].width
        copy_cell(cell, tar_sheet, tar_row_num, key)


def split_sheet(sheet: Worksheet,
                col: Union[str, int],
                save_path: Union[str, Path],
                begin_row: int = 2,
                keys: Union[str, tuple, list, set] = None) -> None:
    """按照某列内容把xlsx文件拆分成多个文件    \n
    :param sheet: 要拆分的工作表
    :param col: 数据列名或列号
    :param begin_row: 数据行开始行号
    :param save_path: 保存路径
    :param keys: 指定只拆分哪些关键字
    :return: None
    """
    Path(save_path).mkdir(parents=True, exist_ok=True)

    if isinstance(keys, str):
        keys = {keys}

    if keys is not None:
        keys = set(keys)

    if isinstance(col, str):
        col = column_index_from_string(col)

    col -= 1

    if begin_row > 1:
        title_rows = sheet[1:begin_row - 1] if begin_row > 2 else (sheet[1],)
    else:
        begin_row = 1
        title_rows = ()

    rows = list(sheet[begin_row:sheet.max_row])
    rows.sort(key=lambda x: str(x[col].value))

    wb = ws = value = None
    for row in rows:
        now_value = '' if row[col].value is None else str(row[col].value)

        if keys is not None and now_value not in keys:
            continue

        # 如果内容和前一个不同，说明前一个已经采集完，可保存并新建工作簿
        if now_value != value:
            if ws is not None:
                wb.save(f'{save_path}\\{make_valid_name(value)}.xlsx')
                wb.close()

            if keys is not None and ws is not None:
                keys.remove(value)
                if not keys:  # 如果keys内没有内容，说明已经采集完毕，退出循环
                    wb = None
                    break

            wb = Workbook()
            ws = wb.active
            for title in title_rows:
                copy_row(title, ws)

        copy_row(row, ws)
        value = now_value

    if wb is not None:
        wb.save(f'{save_path}\\{make_valid_name(value)}.xlsx')
        wb.close()


def find_in_row(row_or_num: Union[tuple, int],
                key: str,
                sheet: Worksheet = None,
                fuzzy: bool = False,
                return_all: bool = False,
                begin_col: Union[int, str] = 1) -> Union[None, tuple]:
    """在某行中查找一个值，返回单元格坐标或坐标组成的元组             \n
    :param row_or_num: 行元组或行号，输入行号时必须输入sheet
    :param key: 要查找的关键字
    :param sheet: 工作表
    :param fuzzy: 是否模糊查找
    :param return_all: 是否返回全部结果，是则以元组返回多个坐标
    :param begin_col: 开始搜索的列号
    :return: 返回单元格坐标，找不到返回None或空元组
    """
    if isinstance(row_or_num, int):
        if not sheet:
            raise ValueError('用行号定位时必须输入sheet')
        row_or_num = sheet[row_or_num]
    elif not isinstance(row_or_num, tuple):
        raise TypeError('行格式不正确')

    if isinstance(begin_col, str):
        begin_col = column_index_from_string(begin_col)

    return _find_by_fuzzy(row_or_num, key, return_all, begin_col) \
        if fuzzy else _find_by_exact(row_or_num, key, return_all, begin_col)


def find_in_col(col_or_num: Union[tuple, int, str],
                key: str,
                sheet: Worksheet = None,
                fuzzy: bool = True,
                return_all: bool = False,
                begin_row: int = 1) -> Union[None, tuple]:
    """在某列中查找一个值，返回单元格坐标或坐标组成的元组             \n
    :param col_or_num: 列元组或列号
    :param key: 要查找的关键字
    :param sheet: 工作表
    :param fuzzy: 是否模糊查找
    :param return_all: 是否返回全部结果，是则以元组返回多个坐标
    :param begin_row: 开始搜索的行号
    :return: 返回单元格坐标，找不到返回None或空元组
    """
    if isinstance(col_or_num, (str, int)) and not sheet:
        raise ValueError('用行号定位时必须输入sheet')

    if isinstance(col_or_num, int):
        col_or_num = sheet[get_column_letter(col_or_num)]
    elif isinstance(col_or_num, str):
        col_or_num = sheet[col_or_num]
    elif not isinstance(col_or_num, tuple):
        raise TypeError('列格式不正确')

    return _find_by_fuzzy(col_or_num, key, return_all, begin_row) \
        if fuzzy else _find_by_exact(col_or_num, key, return_all, begin_row)


def find_in_sheet(sheet: Worksheet,
                  key: str,
                  fuzzy: bool = True,
                  return_all: bool = False) -> Union[None, tuple]:
    """在工作表中查找一个值，返回单元格坐标或坐标组成的元组             \n
    :param sheet: 工作表
    :param key: 要查找的关键字
    :param fuzzy: 是否模糊查找
    :param return_all: 是否返回全部结果，是则以元组返回多个坐标
    :return: 返回单元格坐标，找不到返回None或空元组
    """
    if return_all:
        locs = []
        for row in sheet.rows:
            loc = _find_by_fuzzy(row, key, True) if fuzzy else _find_by_exact(row, key, True)
            locs.extend(loc)
        return tuple(locs)

    else:
        for row in sheet.rows:
            loc = _find_by_fuzzy(row, key) if fuzzy else _find_by_exact(row, key)
            if loc:
                return loc


def _find_by_fuzzy(row_or_col: tuple,
                   key: str,
                   return_all: bool = False,
                   begin: int = 1) -> Union[tuple, None]:
    """在行或列中以模糊查找的方式查找一个值，返回单元格坐标         \n
    :param row_or_col: 行或列元组
    :param key: 要查找的关键字
    :param return_all: 是否返回全部结果，是则以元组返回多个坐标
    :return: 单元格坐标，找不到返回None
    """
    if begin < 1:
        begin = 1

    if return_all:
        return tuple((i.row, i.column) for i in row_or_col[begin - 1:] if key in str(i.value))
    else:
        for i in row_or_col[begin - 1:]:
            if key in str(i.value):
                return i.row, i.column


def _find_by_exact(row_or_col: tuple,
                   key: str,
                   return_all: bool = False,
                   begin: int = 1) -> Union[tuple, None]:
    """在行或列中以精确查找的方式查找一个值，返回单元格坐标  \n
    :param row_or_col: 行或列元组
    :param key: 要查找的关键字
    :return: 单元格坐标，找不到返回None
    """
    if begin < 1:
        begin = 1

    if return_all:
        return tuple((i.row, i.column) for i in row_or_col[begin - 1:] if key == str(i.value))
    else:
        for i in row_or_col[begin - 1:]:
            if key == str(i.value):
                return i.row, i.column
