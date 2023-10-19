import os
import platform
import subprocess
import typing
from pathlib import Path

import numpy as np
import pint
from PySide6.QtWidgets import QMessageBox
import openpyxl.styles as _s
from openpyxl.styles import NamedStyle, Font, PatternFill

from maphis.common.label_hierarchy import LabelHierarchy
from maphis.common.label_image import PropertyType, RegionProperty
from maphis.common.plugin import ActionContext
from maphis.measurement.values import MatrixValue, VectorValue, ureg, ValueType


def get_default_unit(unit: pint.Unit) -> pint.Unit:
    if 'dimensionless' in unit._units:
        return unit
    if 'pixel' in unit._units:
        return pint.Quantity(1, unit).to(ureg['pixel'] ** (unit.dimensionality['[printing_unit]']))._units
    return pint.Quantity(1, unit).to(ureg['millimeter'] ** (unit.dimensionality['[length]']))._units



PyxlCellStyle = typing.Union[
    _s.PatternFill,
    _s.Font,
    _s.Color,
]


cell_styles = {
    'missing_scale': NamedStyle('missing_scale', font=Font(color='9C5700'), fill=PatternFill('solid', fgColor='FFEB9C')),
    'bold_font': NamedStyle('bold_font', font=Font(bold=True)),
    'missing_value': NamedStyle('missing_value', font=Font(color='9C0006'), fill=PatternFill('solid', fgColor='FFC7CE'))
}


scale_flag = {
    'missing': 'px',
    'present': 'mm'
}


class StyledCells:
    def __init__(self, style: NamedStyle):
        self.cells: typing.List[typing.Tuple[int, int]] = []
        self.style: NamedStyle = style

    def add_cells(self, cells: typing.List[typing.Tuple[int, int]]):
        self.cells.extend(cells)


def index_to_column_address(index: int) -> str:
    address: str = ''
    address_length = index // 26 + 1
    for i in range(address_length):
        letter_num = index % 26
        letter = chr(letter_num + 65)
        address += letter
        index = index - letter_num
    return address[::-1]


def is_real_unit(unit: pint.Unit) -> bool:
    # if isinstance(unit, Unit):
    #     return unit.base_unit != BaseUnit.px
    # return all([u.base_unit != BaseUnit.px for u in unit.numerator]) and all([u.base_unit != BaseUnit.px for u in unit.denominator])
    return 'pixel' not in unit


def _ndarray_export_routine(prop_list: typing.List[typing.Tuple[int, str, str, str]], append_row, context: ActionContext) -> typing.List[StyledCells]:
    row = []
    missing_scale_styled_cells: StyledCells = StyledCells(cell_styles['missing_scale'])
    styled_header_cells: StyledCells = StyledCells(cell_styles['bold_font'])
    missing_value_cells: StyledCells = StyledCells(cell_styles['missing_value'])

    prop_list = list(sorted(prop_list, key=lambda tup: tup[0]))
    row_counter = 0

    for i in range(context.storage.image_count):
        photo = context.storage.get_photo_by_idx(i, load_image=False)
        # lab_img = photo['Labels']
        lab_img = photo[context.current_label_name]
        row = [photo.image_name]

        styled_header_cells.add_cells([(row_counter + 1, 1)])

        if photo.image_scale is None:
            missing_scale_styled_cells.add_cells([(row_counter + 1, 1)])

        for label, prop_comp_key, local_key, prop_name in prop_list:
            prop_key = f'{prop_comp_key}.{local_key}'
            if lab_img.get_region_props(label) is None:
                continue
            reg_props = lab_img.get_region_props(label)

            if prop_key not in reg_props:
                continue
            reg_prop = reg_props[prop_key]
            prop: MatrixValue = reg_prop.value
            region_name = lab_img.label_hierarchy[label].name

            for j in range(prop.channel_count):
                if j == 0:
                    row.append(f'{region_name}:{reg_prop.info.name} {prop.channel_names[j]}')
                else:
                    row = ['', f'{region_name}:{reg_prop.info.name} {prop.channel_names[j]}']
                for col in prop.column_names:
                    row.append(col)
                append_row(row)
                row_counter += 1
                # matrix2d: np.ndarray = prop.value[0][j]
                matrix2d: np.ndarray = prop.raw_value[j]
                for r in range(matrix2d.shape[0]):
                    if j == 0 and r == 0:
                        row = [scale_flag['present'] if photo.image_scale is not None else scale_flag['missing'], prop.row_names[r]]
                        if photo.image_scale is None:
                            missing_scale_styled_cells.add_cells([(row_counter + 1, 1)])
                    else:
                        row = ['', prop.row_names[r]]
                    for c in range(matrix2d.shape[1]):
                        if np.isnan(matrix2d[r, c]):
                            row.append('N/A')
                            missing_value_cells.add_cells([(row_counter + 1, 2 + c + 1)])
                        else:
                            row.append(matrix2d[r, c])
                    append_row(row)
                    row_counter += 1
                row = ['']
                append_row(row)
                row_counter += 1
            append_row([''])
            row_counter += 1
    return [missing_scale_styled_cells, styled_header_cells, missing_value_cells]


def _filter_by_NDArray(prop_tuple_list: typing.List[typing.Tuple[int, str, str, str]], context: ActionContext) -> \
        typing.Tuple[typing.List[typing.Tuple[int, str, str, str]], typing.List[typing.Tuple[int, str, str, str]]]:
    ndarray_props: typing.List[typing.Tuple[int, str, str, str]] = []
    other_props: typing.List[typing.Tuple[int, str, str, str]] = []

    for label, prop_comp_key, local_key, prop_name in prop_tuple_list:
        # dot_split = prop_key.split('.')
        prop_key = f'{prop_comp_key}.{local_key}'
        # comp_key = '.'.join(dot_split[:-1])
        # computation = self.computation_widget.computations_model.computations_dict[comp_key]
        if prop_comp_key not in context.property_computations: #self.computation_widget.computations_model.computations_dict:
            continue
        # computation = self.computation_widget.computations_model.computations_dict[prop_key]
        computation = context.property_computations[prop_comp_key]
        if computation.example(local_key).value.value_type == ValueType.Matrix:
            ndarray_props.append((label, prop_comp_key, local_key, prop_name))
        else:
            other_props.append((label, prop_comp_key, local_key, prop_name))
    return ndarray_props, other_props


def get_prop_tuple_list(context: ActionContext) -> typing.List[typing.Tuple[int, str, str, str]]:
    Label = int
    PropCompKey = str
    PropName = str
    PropKey = str

    prop_tuple_set: typing.Set[typing.Tuple[Label, PropCompKey, PropKey, PropName]] = set()
    for i in range(context.storage.image_count):
        photo = context.storage.get_photo_by_idx(i, load_image=False)
        for prop in photo[context.current_label_name].prop_list:
            #prop_tuple = (prop.info.key, prop.label, prop.info.name)
            prop_tuple = (prop.label, prop.prop_comp_key, prop.local_key, prop.info.name)
            prop_tuple_set.add(prop_tuple)
    return list(sorted(prop_tuple_set, key=lambda tup: tup[:3]))


def _group_measurements_by_sheet(prop_tup_list: typing.List[typing.Tuple[int, str, str, str]], context: ActionContext) \
        -> typing.Dict[str, typing.List[typing.Tuple[int, str, str, str]]]:
    sheet_grouped_props: typing.Dict[str, typing.List[typing.Tuple[int, str, str, str]]] = {}

    for label, prop_comp_key, local_key, prop_name in prop_tup_list:
        # dot_split = prop_key.split('.')
        # comp_key = '.'.join(dot_split[:-1])
        # computation = self.computation_widget.computations_model.computations_dict[prop_key]
        # prop_key = f'{prop_comp_key}.{local_key}'
        computation = context.property_computations[prop_comp_key]
        # prop_key_local = dot_split[-1]
        sheet_grouped_props.setdefault(computation.target_worksheet(local_key), []).append((label,
                                                                                                 prop_comp_key,
                                                                                                 local_key,
                                                                                                 prop_name))

    return sheet_grouped_props


def _tabular_export_routine(prop_tuple_list: typing.List[typing.Tuple[int, str, str, str]],
                            append_row: typing.Callable[[typing.List[typing.Any]], None],
                            context: ActionContext) -> typing.List[StyledCells]:
    example_props: typing.Dict[str, RegionProperty] = {}
    column_names = ['photo_name', 'unit']
    missing_scale_cells = StyledCells(cell_styles['missing_scale'])
    header_cells = StyledCells(cell_styles['bold_font'])
    missing_value_cells = StyledCells(cell_styles['missing_value'])
    # style.add_style('pattern_fill', _s.PatternFill(patternType='solid', fgColor='F5EA25'))

    lab_hier: LabelHierarchy = context.storage.get_label_hierarchy(context.current_label_name)
    for label, prop_comp_key, local_key, prop_name in prop_tuple_list:
        prop_key = f'{prop_comp_key}.{local_key}'
        # dot_split = prop_key.split('.')
        # comp_key = '.'.join(dot_split[:-1])
        # computation = self.computation_widget.computations_model.computations_dict[prop_key]
        computation = context.property_computations[prop_comp_key]
        prop = computation.example(local_key)
        example_props[prop_key] = prop
        # if prop.prop_type == PropertyType.Vector or prop.num_vals > 1:
        if prop.value.count > 1:
            prop_val: VectorValue = prop.value
            if len(prop_val.column_names) != prop_val.count:
                col_names = [str(i) for i in range(prop_val.count)]
            else:
                col_names = prop_val.column_names
            # unit_suffix = f'[{prop.value.unit}]' if prop.value.unit.base_unit != BaseUnit.none else ''
            unit_suffix = f'{prop_val.value:~P}'
            for i in range(prop_val.count):
                #column_names.append(f'{prop.info.key}_{prop.val_names[i]}:{self.state.colormap.label_names[label]}')
                # column_names.append(f'{prop.info.key}_{prop.val_names[i]}:{self.state.label_hierarchy.nodes[label].name}')
                column_names.append(f'{lab_hier[label].name}:{prop.info.name}_{col_names[i]} {unit_suffix}')
        else:
            #column_names.append(f'{prop.info.key}:{self.state.colormap.label_names[label]}')
            # column_names.append(f'{prop.info.key}:{self.state.label_hierarchy.nodes[label].name}')
            # unit_str = str(context.units.get_default_unit(prop.value.unit))
            unit_str = get_default_unit(prop.value.unit)
            unit_specifier = f'[{unit_str}]' if len(unit_str) > 0 else ''
            column_names.append(f'{lab_hier[label].name}:{prop.info.name} {unit_specifier}')
    header_cells.add_cells([(1, c + 1) for c in range(len(column_names))])

    append_row(column_names)
    for i in range(context.storage.image_count):
        photo = context.storage.get_photo_by_idx(i, load_image=False)
        if photo.image_scale is None:
            missing_scale_cells.add_cells([(i + 2, 2)])
        row = [photo.image_name, scale_flag['present'] if photo.image_scale is not None else scale_flag['missing']]
        # label_img = photo['Labels']
        label_img = photo[context.current_label_name]
        col = 0
        for label, prop_comp_key, local_key, prop_name in prop_tuple_list:
            prop_key = f'{prop_comp_key}.{local_key}'
            if label_img.get_region_props(
                    label) is None:  # So this LabelImg does not have props for `label`, insert a sequence of 'missing value' symbols (-1)
                ex_prop = example_props[prop_key]
                for _ in range(ex_prop.value.count):
                    row.append('N/A')
                    missing_value_cells.add_cells([(i + 2, col + 2 + 1)])
                    col += 1
            else:
                reg_props = label_img.get_region_props(label)
                if prop_key not in reg_props:  # This label region does not have property with `prop_key`, insert a sequence of `missing value` symbols (-1)
                    ex_prop = example_props[prop_key]
                    for _ in range(ex_prop.value.count):
                        row.append('N/A')
                        missing_value_cells.add_cells([(i + 2, col + 2 + 1)])
                        col += 1
                else:  # Finally, insert actual values for present property
                    reg_prop: RegionProperty = reg_props[prop_key]
                    if reg_prop.value.count > 1:
                        prop_val: VectorValue = reg_prop.value
                        unit_: pint.Unit = prop_val.unit
                        targ_unit = get_default_unit(unit_)
                        # mult = (10 ** (int(reg_prop.value[1].prefix) - int(targ_unit.prefix))) ** reg_prop.value[1].dim
                        # val_ = Value(reg_prop.value[0][0], unit_)
                        val_ = prop_val.value.to(targ_unit)
                        for val in val_.magnitude:
                            # val_.value = val
                            # n_val = convert_value(val_, targ_unit)
                            row.append(val)
                            if not is_real_unit(targ_unit):
                                missing_scale_cells.add_cells([(i + 2, col + 2 + 1)])
                            col += 1
                            # break  # TODO how to handle exporting vector of values to CSV?
                    else:
                        # targ_unit = context.units.get_default_unit(reg_prop.value.unit)
                        targ_unit = get_default_unit(reg_prop.value.unit)
                        # conv_val = convert_value(reg_prop.value, targ_unit)
                        conv_val = reg_prop.value.value.to(targ_unit)
                        row.append(conv_val.magnitude)
                        # if targ_unit.base_unit == unit_store.units['px'].base_unit:
                        if not is_real_unit(targ_unit):
                            missing_scale_cells.add_cells([(i + 2, col + 2 + 1)])
                        col += 1
        append_row(row)

    return [missing_scale_cells, header_cells, missing_value_cells]


def open_project_folder_in_explorer(context: ActionContext):
    if context.storage is None:  # Shouldn't even be necessary, as the QAction shouldn't be enabled in that case
        return
    if platform.system() == "Windows":
        os.startfile(context.storage.location)
    elif platform.system() == "Darwin":
        subprocess.Popen(["open", str(context.storage.location)])
    else:
        subprocess.Popen(["xdg-open", str(context.storage.location)])


def show_export_success_message(folder: Path, filenames: typing.List[str], context: ActionContext):
    filenames_in_rows = '\n'.join(filenames)
    if QMessageBox.information(None, 'Export finished',
                               f'The results were saved in the folder {folder} as files:\n{filenames_in_rows}\nDo you want to open the folder?',
                               QMessageBox.Yes | QMessageBox.No, QMessageBox.No) == QMessageBox.Yes:
        open_project_folder_in_explorer(context)
