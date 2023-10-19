import csv
from pathlib import Path
import typing
from typing import Optional

import openpyxl
import openpyxl.worksheet.worksheet as worksheet
from openpyxl.styles import Color, PatternFill, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import CellIsRule, Rule
from PySide6.QtWidgets import QApplication

from maphis.common.action import action_info, GeneralActionContext
from maphis.common.utils import attempt_to_save_with_retries
from maphis.common.common import Info
from maphis.common.plugin import GeneralAction, ActionContext
from maphis.common.state import State
from maphis.plugins.maphis.general.common import _ndarray_export_routine, _filter_by_NDArray, \
    _group_measurements_by_sheet, _tabular_export_routine, show_export_success_message, get_prop_tuple_list, StyledCells
from maphis.project.annotation import KeypointAnnotation, AnnotationType
from maphis.project.project import Project
from maphis.tools.landmarks import get_landmark_in_real_units


@action_info('Export to XLSX', 'Export measurements to a XLSX file', 'Export:Measurements')
class ExportXLSX(GeneralAction):

    def __init__(self, info: Optional[Info] = None):
        super().__init__(info)

    def __call__(self, state: State, context: ActionContext) -> None:
        wb = openpyxl.Workbook()
        prop_tuple_list = get_prop_tuple_list(context)
        # self._tabular_export_routine(ws.append)
        nd_props, other_props = _filter_by_NDArray(prop_tuple_list, context)

        sheet_grouped_props = _group_measurements_by_sheet(other_props, context)

        for sheet_name, prop_list in sheet_grouped_props.items():
            ws: worksheet.Worksheet = wb.create_sheet(sheet_name)
            styles = _tabular_export_routine(prop_list, ws.append, context)
            for style in styles:
                for cell in style.cells:
                    ws.cell(*cell).style = style.style

        sheet_nd_props = _group_measurements_by_sheet(nd_props, context)

        for sheet_name, prop_list in sheet_nd_props.items():
            ws = wb.create_sheet(sheet_name)
            cell_styles: typing.List[StyledCells] = _ndarray_export_routine(prop_list, ws.append, context)
            for styled_cells in cell_styles:
                for cell in styled_cells.cells:
                    ws.cell(*cell).style = styled_cells.style

        self._write_landmarks_to_workbook(wb, state.project)

        if 'Sheet' in wb:
            wb.remove(wb['Sheet'])

        # wb.save(str(context.storage.location / f'{context.current_label_name}_results.xlsx'))

        final_path, _ = attempt_to_save_with_retries(_save, 'Save as...', 'single_file', ['xlsx'], QApplication.activeWindow(),
                                     path=context.storage.location / f'{context.current_label_name}_results.xlsx',
                                     object=wb)
        if final_path is not None:
            show_export_success_message(final_path.parent, [final_path.name], context)

        # show_export_success_message(context.storage.location, [f'{context.current_label_name}_results.xlsx'], context)

    @property
    def general_action_context(self) -> GeneralActionContext:
        return GeneralActionContext.Measurements

    def _write_landmarks_to_workbook(self, wb: openpyxl.Workbook, project: Project):
        ws = wb.create_sheet('Landmarks')

        ws.append(['image_name', 'landmark_name', 'x (px)', 'y (px)', 'x (mm)', 'y (mm)'])

        for photo in project.storage.images:
            lm_annotations: typing.List[KeypointAnnotation] = photo.get_annotations(AnnotationType.Keypoint)
            if len(lm_annotations) == 0:
                continue
            for annotation in lm_annotations:
                for kp in annotation.kps:
                    real_kp = get_landmark_in_real_units(kp, photo)
                    ws.append((photo.image_name, kp.name, kp.x, kp.y, real_kp.x, real_kp.y))


    # @property
    # def group(self) -> str:
    #     return 'export'


def _save(path: Path, object: openpyxl.Workbook):
    object.save(path)
