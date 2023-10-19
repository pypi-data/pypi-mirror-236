import dataclasses
import itertools
import re
import typing
from pathlib import Path
from typing import Optional
import importlib.resources

import numpy as np
import openpyxl
import pint
import pyparsing
from PySide6.QtCore import QUrl
from PySide6.QtWidgets import QMessageBox, QWidget, QPushButton, QVBoxLayout, QDialogButtonBox, \
    QDialog, QApplication, QScrollArea, QSizePolicy, QHBoxLayout, QFileDialog

from maphis import MAPHIS_PATH
from maphis.common.action import action_info, GeneralActionContext
from maphis.common.common import Info
from maphis.common.photo import Photo
from maphis.common.plugin import GeneralAction, ActionContext
from maphis.common.state import State
from maphis.common.storage import Storage
from maphis.common.utils import open_with_default_app, let_user_save_path, attempt_to_save_with_retries, save_text, \
    catch_exception_and_show_messagebox
from maphis.measurement.region_property import RegionProperty
from maphis.measurement.values import ureg, VectorValue
from maphis.plugins.profile_register.general.profiles import get_median_profile, merge_profiles
from maphis.plugins.profile_register.general.tag_mapping_widget import TagMappingWidget
from maphis.project.project import Project
from maphis.tags.tag_chooser import Tags
from maphis.tags.tags_choosers import GroupsOfTags


BODY = 16842752
PROFILE_KEY = "maphis.plugins.profile_register.properties.profile.profile"
TAG_REGEX = '\w+\D+\d*'
TAG_LIST_REGEX = f'\[({TAG_REGEX}\s*,\s*)*({TAG_REGEX})\]'
REGISTER_CONF_LINE_REGEX = re.compile(f'^({TAG_LIST_REGEX})\s*(\[({TAG_LIST_REGEX}\s*,\s*)*({TAG_LIST_REGEX})\])')

BR_L = pyparsing.Suppress(pyparsing.Literal('['))
BR_R = pyparsing.Suppress(pyparsing.Literal(']'))
TAG = pyparsing.Word(pyparsing.alphas, pyparsing.alphanums + '_')
TAG_LIST = pyparsing.Group(BR_L + pyparsing.delimited_list(TAG, delim=',') + BR_R)
CONF_LINE = TAG_LIST + BR_L + pyparsing.Group(pyparsing.delimited_list(TAG_LIST, delim=',')) + BR_R


@dataclasses.dataclass
class SettingWidget:
    toplevel_dialog: typing.Optional[QDialog]
    scroll_area: typing.Optional[QScrollArea]
    main_widget: typing.Optional[QWidget]
    main_layout: typing.Optional[QVBoxLayout]
    tag_mapping_layout: typing.Optional[QVBoxLayout]
    btnAdd: typing.Optional[QPushButton]
    btnSaveConf: typing.Optional[QPushButton]
    btnLoadConf: typing.Optional[QPushButton]
    diagButtonBox: typing.Optional[QDialogButtonBox]


@dataclasses.dataclass
class ProfileFusionResult:
    src_median_profile: np.ndarray
    dst_median_profile: np.ndarray
    aligned_profiles: np.ndarray


@action_info(name='Profile fusion',
             description='Fuse body profiles based on their tags. '
                         'A median profile is created for each group of images that matches the given tags.',
             group='Plugins')
class ProfileFusion(GeneralAction):
    def __init__(self, info: Optional[Info] = None):
        super().__init__(info)
        # self._setting_widget: QDialog = self._build_settings_widget()
        self._setting_widget: typing.Optional[SettingWidget] = None
        self._state: typing.Optional[State] = None

    def __call__(self, state: State, _: ActionContext):
        if not self._check_for_profile_measurement(state):
            return
        self._state = state

        self._setting_widget = self._build_settings_widget2(state, state.storage, state.project)
        self._setting_widget.toplevel_dialog.open()

    @property
    def general_action_context(self) -> GeneralActionContext:
        return GeneralActionContext.Measurements

    def _get_median_profile_for(self, iteration_tag: str, group_tags: typing.Set[str], state: State) -> typing.Optional[np.ndarray]:
        tags_set = set(group_tags).union({iteration_tag})

        photos = state.storage.photos_satisfying_tags(tags_set)

        if len(photos) == 0:
            return None
        profiles: np.ndarray = np.zeros((len(photos), 40), np.float32)
        print(f'number of photos satisfying the tag set {tags_set} = {len(photos)}')
        for i, photo in enumerate(photos):
            profiles[i] = np.array(photo['Labels'].region_props[BODY][PROFILE_KEY].value[0])
        median = get_median_profile(profiles, show_fig=False)
        return median

    def _check_for_profile_measurement(self, state: State) -> bool:
        photos_missing_profiles: typing.List[Photo] = []
        for photo in state.storage.images:
            if BODY not in photo['Labels'].region_props or PROFILE_KEY not in photo['Labels'].region_props[BODY]:
                photos_missing_profiles.append(photo)
        if len(photos_missing_profiles) > 0:
            QMessageBox.information(None, 'Missing profile measurements', f'{len(photos_missing_profiles)} photo(s) do(es) not have body profile measurement. Please compute them first.')
            return False
        return True

    def _execute(self, state: State, _: ActionContext):
        output_folder = Path(state.storage.location) / 'registered_profiles'
        if not output_folder.exists():
            output_folder.mkdir(exist_ok=True)
        # TODO add an option to overwrite the existing results
        # if self.chk_delete_existing.isChecked():
        #     os.remove(output_folder / 'profiles.xlsx')
        if (worksheet_path := output_folder / 'profiles.xlsx').exists():
            wb = openpyxl.load_workbook(worksheet_path, read_only=False)
        else:
            wb = openpyxl.Workbook()
            ws_med_prof = wb.active
            ws_med_prof.title = 'Median profiles'
            ws_med_prof.append(['ProfileID', 'unit'] + [f'G_BP_{i}' for i in range(40)])

            ws_aligned = wb.create_sheet('Aligned profiles')
            ws_aligned.append(['ProfileID', 'AlignedTo', 'unit'] + [f'G_BP_{i}' for i in range(40)])

        ws_med_prof = wb['Median profiles']
        ws_aligned = wb['Aligned profiles']

        mappings = self._gather_mappings()
        for mapping in mappings:
            src, dsts = mapping
            fusion_result_mm, ignored_photos = self._register(src, dsts, state)
            src_median_profile_id = ','.join(src)
            dst_median_profile_id = '_'.join([','.join(dst_tags) for dst_tags in dsts])
            if fusion_result_mm is not None:
                ws_med_prof.append([src_median_profile_id, 'mm'] + [float(val) for val in fusion_result_mm.src_median_profile])
                ws_med_prof.append([dst_median_profile_id, 'mm'] + [float(val) for val in fusion_result_mm.dst_median_profile])
                ws_aligned.append([src_median_profile_id, dst_median_profile_id, 'mm'] + [float(val) for val in fusion_result_mm.aligned_profiles])
            if len(ignored_photos) > 0:
                QMessageBox.information(self._setting_widget.toplevel_dialog, 'Some profiles were ignored',
                                               f'{len(ignored_photos)} profile(s) were ignored because they were measured in pixel units.',
                                               QMessageBox.StandardButton.Ok)
        self._attempt_to_save_results(wb, worksheet_path)

    def _attempt_to_save_results(self, wb: openpyxl.Workbook, path: Path) -> bool:
        results_saved_or_cancelled = False
        while not results_saved_or_cancelled:
            try:
                wb.save(path)
                wb.close()
                results_saved_or_cancelled = True
                self._show_success_info_dialog(path)
                return True
            except PermissionError:
                path = self._show_error_and_ask_for_other_path(path)
                if path is None:
                    return False
        return False

    def _show_error_and_ask_for_other_path(self, old_path: Path) -> typing.Optional[Path]:
        diag = QMessageBox(self._setting_widget.toplevel_dialog)
        diag.setWindowTitle('Cannot save the results')
        diag.setText(
            f"The file {old_path} cannot be used for saving (maybe it's opened in another program).")
        diag.addButton(QMessageBox.StandardButton.Cancel)
        diag.addButton(QMessageBox.StandardButton.Open)
        button: QPushButton = diag.button(QMessageBox.StandardButton.Open)
        button.clicked.connect(diag.accept)
        button.setText('Save as...')

        if diag.exec_() == QMessageBox.DialogCode.Accepted:
            return let_user_save_path(old_path.parent, 'Save as', 'single_file', ['xlsx'], self._setting_widget.toplevel_dialog)
        return None

    def _gather_mappings(self) -> typing.List[typing.Tuple[Tags, GroupsOfTags]]:
        mappings: typing.List[typing.Tuple[Tags, GroupsOfTags]] = []
        for i in range(self._setting_widget.tag_mapping_layout.count()):
            tag_mapping: TagMappingWidget = self._setting_widget.tag_mapping_layout.itemAt(i).widget()
            mappings.append(tag_mapping.get_mapping())
        return mappings

    def _register(self, src: typing.List[str], dsts: typing.List[typing.List[str]], state: State) -> typing.Tuple[typing.Optional[ProfileFusionResult], typing.List[Photo]]:
        src_profiles_mm, ignored_photos = self._gather_profiles(src, state)
        ignored_photos: typing.Set[Photo] = set(ignored_photos)

        if src_profiles_mm is None:
            return None, list(ignored_photos)
        src_median_profile_mm = get_median_profile(src_profiles_mm)

        dst_median_profiles_mm: typing.List[np.ndarray] = []
        for dst in dsts:
            dst_profiles_mm, dst_ignored_photos = self._gather_profiles(dst, state)
            ignored_photos.update(dst_ignored_photos)
            if dst_profiles_mm is not None:
                dst_median_profiles_mm.append(get_median_profile(dst_profiles_mm))

        if len(dst_median_profiles_mm) > 0:
            dst_median_profile_mm = get_median_profile(np.array(dst_median_profiles_mm))
        else:
            dst_median_profile_mm = None

        if dst_median_profile_mm is not None:
            aligned_profiles_mm = merge_profiles(src_median_profile_mm, dst_median_profile_mm, x_weight=0.5, y_weight=0.5)
            fusion_result_mm = ProfileFusionResult(src_median_profile_mm, dst_median_profile_mm, aligned_profiles_mm)
        else:
            fusion_result_mm = None
        return fusion_result_mm, list(ignored_photos)

    def _gather_profiles(self, tags: typing.List[str], state: State) \
            -> typing.Tuple[Optional[np.ndarray], typing.List[Photo]]:
        photos: typing.List[Photo] = state.storage.photos_satisfying_tags(set(tags))

        photos_with_scale: typing.List[Photo] = [photo for photo in photos if photo.image_scale is not None]
        photos_no_scale: typing.List[Photo] = [photo for photo in photos if photo.image_scale is None]

        # profiles_real_units: typing.List[np.ndarray] = [np.array(photo['Labels'].region_props[BODY][PROFILE_KEY].value[0]) for photo in photos_with_scale]

        profiles_mm: typing.List[np.ndarray] = []
        for photo in photos_with_scale:
            profile_prop: RegionProperty = photo['Labels'].region_props[BODY][PROFILE_KEY]
            profile_vector: VectorValue = profile_prop.value
            profile_mm = np.zeros((profile_vector.count, ))
            converted: pint.Quantity = profile_prop.value.value.to(ureg['millimeter'])
            for idx, val in enumerate(converted.magnitude):
                # value = Value(val, profile_prop.value[1])
                # value_mm = convert_value(value, state.units.units['mm'])
                profile_mm[idx] = val
            profiles_mm.append(profile_mm)

        profiles_mm: typing.Optional[np.ndarray] = np.array(profiles_mm) if len(profiles_mm) > 0 else None

        return profiles_mm, photos_no_scale

    def _show_success_info_dialog(self, output_path: Path):
        diag_box = QMessageBox(QMessageBox.Icon.Information, "Profile fusion completed",
                               f"The results are stored in {output_path}. Do you wish to open the folder?",
                               QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, QApplication.activeWindow())
        diag_box.button(QMessageBox.StandardButton.Yes).clicked.connect(diag_box.accept)
        diag_box.button(QMessageBox.StandardButton.No).clicked.connect(diag_box.reject)

        if diag_box.exec_() == QMessageBox.DialogCode.Accepted:
            open_with_default_app(output_path.parent)
        diag_box.close()
        diag_box.deleteLater()

    def setting_widget(self) -> typing.Optional[QWidget]:
        return None

    def _remove_tag_mapping(self, tag_map_widget: TagMappingWidget):
        if self._setting_widget.tag_mapping_layout.count() == 1:
            return
        tag_map_widget.hide()
        self._setting_widget.tag_mapping_layout.removeWidget(tag_map_widget)
        tag_map_widget.deleteLater()
        last_mapping: TagMappingWidget = self._setting_widget.tag_mapping_layout.itemAt(self._setting_widget.tag_mapping_layout.count() - 1).widget()
        last_mapping.enable_remove_button(self._setting_widget.tag_mapping_layout.count() > 1)
        self._update_apply_button()

    def _add_tag_mapping(self, state: State, storage: Storage):
        new_mapping = TagMappingWidget(state, parent=self._setting_widget.main_widget)
        new_mapping.remove_this.connect(self._remove_tag_mapping)
        new_mapping.complete_status_changed.connect(lambda _: self._update_apply_button())
        self._setting_widget.tag_mapping_layout.addWidget(new_mapping)

        for idx in range(self._setting_widget.tag_mapping_layout.count()):
            tag_map_widget: TagMappingWidget = self._setting_widget.tag_mapping_layout.itemAt(idx).widget()
            tag_map_widget.enable_remove_button(self._setting_widget.tag_mapping_layout.count() > 1)
        self._update_apply_button()

    def _update_apply_button(self):
        all_mappings_complete = True
        for i in range(self._setting_widget.tag_mapping_layout.count()):
            map_widget: TagMappingWidget = self._setting_widget.tag_mapping_layout.itemAt(i).widget()
            all_mappings_complete = all_mappings_complete and map_widget.is_complete_and_matches_photos
        self._setting_widget.diagButtonBox.button(QDialogButtonBox.StandardButton.Apply).setEnabled(all_mappings_complete)
        self._setting_widget.btnSaveConf.setEnabled(all_mappings_complete)

    def _build_settings_widget2(self, state: State, storage: Storage, project: Project) -> SettingWidget:
        sett_widget = SettingWidget(
            toplevel_dialog=QDialog(),
            scroll_area=QScrollArea(),
            main_widget=QWidget(),
            main_layout=QVBoxLayout(),
            tag_mapping_layout=QVBoxLayout(),
            btnAdd=QPushButton('+'),
            btnSaveConf=QPushButton('Save configuration...'),
            btnLoadConf=QPushButton('Load configuration...'),
            diagButtonBox=QDialogButtonBox(QDialogButtonBox.StandardButton.Apply | QDialogButtonBox.StandardButton.Cancel)
        )
        sett_widget.toplevel_dialog.setWindowTitle('Profile fusion')
        sett_widget.toplevel_dialog.setLayout(QVBoxLayout())
        sett_widget.toplevel_dialog.layout().addWidget(sett_widget.scroll_area)
        sett_widget.toplevel_dialog.setMinimumHeight(600)
        sett_widget.toplevel_dialog.setMinimumWidth(800)

        sett_widget.main_widget.setSizePolicy(QSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Expanding))
        sett_widget.scroll_area.setSizeAdjustPolicy(QScrollArea.SizeAdjustPolicy.AdjustToContents)
        sett_widget.scroll_area.setWidgetResizable(True)

        sett_widget.main_widget.setLayout(sett_widget.main_layout)

        sett_widget.main_layout.addLayout(sett_widget.tag_mapping_layout)
        sett_widget.main_layout.addWidget(sett_widget.btnAdd)
        sett_widget.main_layout.addStretch()

        sett_widget.btnAdd.clicked.connect(lambda: self._add_tag_mapping(state, storage))

        sett_widget.scroll_area.setWidget(sett_widget.main_widget)

        top_level_buttons_layout = QHBoxLayout()
        sett_widget.toplevel_dialog.layout().addLayout(top_level_buttons_layout)

        sett_widget.btnSaveConf.clicked.connect(self._save_configuration)
        sett_widget.btnSaveConf.setEnabled(False)

        sett_widget.btnLoadConf.clicked.connect(self._load_configuration)

        top_level_buttons_layout.addWidget(sett_widget.btnSaveConf)
        top_level_buttons_layout.addWidget(sett_widget.btnLoadConf)

        top_level_buttons_layout.addStretch()

        button_diag_box = sett_widget.diagButtonBox #QDialogButtonBox(QDialogButtonBox.StandardButton.Apply | QDialogButtonBox.StandardButton.Cancel)
        top_level_buttons_layout.addWidget(button_diag_box)
        # button_diag_box.button(QDialogButtonBox.StandardButton.Apply).clicked.connect(sett_widget.toplevel_dialog.accept)
        button_diag_box.button(QDialogButtonBox.StandardButton.Apply).clicked.connect(lambda: self._execute(state, None))
        button_diag_box.button(QDialogButtonBox.StandardButton.Apply).setEnabled(False)
        button_diag_box.button(QDialogButtonBox.StandardButton.Cancel).clicked.connect(sett_widget.toplevel_dialog.close)

        # self._setting_widget.diagButtonBox = button_diag_box

        self._setting_widget = sett_widget

        self._add_tag_mapping(state, storage)

        return sett_widget

    def _load_configuration(self):
        if (recent_folder := self._get_recent_open_save_location()) is None:
            recent_folder = QUrl()
        else:
            recent_folder = QUrl.fromLocalFile(str(recent_folder))
        result = QFileDialog.getOpenFileUrl(self._setting_widget.toplevel_dialog, 'Load configuration from...', dir=recent_folder)
        qurl: QUrl = result[0]
        if qurl.isEmpty():
            return
        load_path = Path(qurl.toLocalFile())
        self._save_recent_open_save_location(load_path.parent)
        self._try_to_load_configuration(load_path=load_path)

    @catch_exception_and_show_messagebox('Loading configuration failed',
                                         'Could not load configuration from {load_path}',
                                         QMessageBox.Icon.Critical, print_exception=True)
    def _try_to_load_configuration(self, load_path: Path):
        with open(load_path, 'r') as f:
            str_mappings = [line.strip() for line in f.readlines()]
        valid_mappings: typing.List[typing.Tuple[Tags, GroupsOfTags]] = []
        invalid_mappings: typing.List[str] = []
        for mapping in str_mappings:
            if not CONF_LINE.matches(mapping):
                invalid_mappings.append(mapping)
            else:
                valid_mappings.append(tuple(CONF_LINE.parse_string(mapping).as_list()))

        if len(invalid_mappings) > 0:
            # TODO show a message box
            return
        how_many_mapping_widgets_to_add = len(valid_mappings) - self._setting_widget.tag_mapping_layout.count()
        for i in range(how_many_mapping_widgets_to_add):
            self._add_tag_mapping(self._state, None)

        for i, mapping in enumerate(valid_mappings):
            tag_map_widg: TagMappingWidget = self._setting_widget.tag_mapping_layout.itemAt(i).widget()
            tag_map_widg.set_mapping(mapping)

    def _save_configuration(self):
        mappings = self._gather_mappings()
        if (recent_folder := self._get_recent_open_save_location()) is None:
            recent_folder = Path(QUrl().toString())
        else:
            recent_folder = Path(QUrl.fromLocalFile(str(recent_folder)).toString())
        save_path = let_user_save_path(recent_folder, extensions=['txt'], parent=QApplication.activeWindow())
        if save_path is None:
            return
        self._save_recent_open_save_location(save_path.parent)

        lines = []
        for tags, groups_of_tags in mappings:
            line = f'{dequote(str(tags))} {dequote(str(groups_of_tags))}\n'
            lines.append(line)

        config = '\n'.join(lines)

        attempt_to_save_with_retries(save_text, 'Save as...', 'single_file',
                                     ['txt'], QApplication.activeWindow(),
                                     text=config, path=save_path)

    def _get_recent_open_save_location(self) -> typing.Optional[Path]:
        try:
            path = MAPHIS_PATH / 'plugins' / 'profile_register' / 'general'
            if not (recent_path_file := path / 'recent_save_open_path.txt').exists():
                return Path.home()
            with open(recent_path_file, 'r') as f:
                recent_path = Path(f.read())
            return recent_path
        except Exception as e:
            print(e)
            # TODO log it
            return None

    def _save_recent_open_save_location(self, loc: Path):
        try:
            # with importlib.resources.path('maphis.plugins.profile_register.general', '.') as path:
            path = MAPHIS_PATH / 'plugins' / 'profile_register' / 'general'
            with open(path / 'recent_save_open_path.txt', 'w') as f:
                f.write(str(loc))
        except Exception as e:
            print(e)
            # TODO log it
            return


def dequote(text: str) -> str:
    return text.replace('"', '').replace("'", '')